"""Smoke test for the closer workbook: recalculates it with LibreOffice and asserts the invariants a reader would trip over.
Usage: check_workbook.py [path]   (exit 1 on any failure)"""
from __future__ import annotations
import os, re, shutil, subprocess, sys, tempfile
from pathlib import Path
from openpyxl import load_workbook
ROOT = Path(__file__).resolve().parents[1]; sys.path.insert(0, str(ROOT))
from anteparo.export.callsheet import STATUSES, BANDS, COLS

path = Path(sys.argv[1] if len(sys.argv) > 1 else ROOT / "data/out/ANTEPARO_CALL_SHEET.xlsx")
tmp = Path(tempfile.mkdtemp(prefix="anteparo_check_"))
prof = tmp / "lo_profile"
r = subprocess.run(["soffice", f"-env:UserInstallation=file://{prof}", "--headless", "--calc", "--convert-to", "xlsx", "--outdir", str(tmp), str(path)], capture_output=True, text=True, timeout=600)
recalc = tmp / path.name
fails = []
def check(cond, msg):
    if not cond: fails.append(msg)
PHONE = re.compile(r"^(\(\d{2}\) \d{4,5}-\d{4}|\d{4}-\d{4}( \(no DDD\))?|\d{5}-\d{4} \(no DDD\)|\(\d{2}\) \d{4}-\d{3} \[digit missing\]|0[3589]00 \d{3} \d{4}|0[3589]00 \d+|\d+ \[check\])$")
if not recalc.exists():
    fails.append(f"LibreOffice could not recalculate: {r.stderr[-300:]}")
else:
    wb = load_workbook(recalc, data_only=True); raw = load_workbook(path)
    errs = [(ws.title, c.coordinate, c.value) for ws in wb.worksheets for row in ws.iter_rows() for c in row if isinstance(c.value, str) and c.value.startswith(("#", "Err:"))]
    check(not errs, f"formula errors: {errs[:5]}")
    asm = wb["Assumptions"]
    check(all(asm[f"A{i}"].value not in ("?", "*", "~") for i in range(8, 13)), "Assumptions band code is an Excel wildcard")
    bands = {b for b, *_ in BANDS}
    total = 0; agio_sum = 0.0; holds = 0; face_hold = 0.0
    for title in ("CALL SHEET", "FINANCIAL CREDITORS", "STATE-OWNED CREDITORS"):
        if title not in wb.sheetnames: continue
        ws = wb[title]; H = [c.value for c in ws[1]]; ix = {h: i for i, h in enumerate(H)}
        check(H == [h for h, _, _ in COLS], f"{title}: header order differs from COLS")
        rows = list(ws.iter_rows(min_row=2, values_only=True)); total += len(rows)
        for k, row in enumerate(rows, start=2):
            g = lambda h: row[ix[h]]
            check(g("DEBTOR BAND") in bands - {"D"}, f"{title} r{k}: band {g('DEBTOR BAND')!r}")
            check(g("Court") != "TJ", f"{title} r{k}: bare 'TJ' court")
            check(isinstance(g("Years to payment"), (int, float)) and g("Years to payment") >= 1.0, f"{title} r{k}: years {g('Years to payment')!r}")
            fund = g("FUND PRICE @ IRR (model, R$)")
            check(isinstance(fund, (int, float)) and 0 <= fund <= g("FACE VALUE (as printed)"), f"{title} r{k}: fund price outside [0, face]")
            if g("PRICING HOLD (why unpriced)"):
                holds += 1; face_hold += float(g("FACE VALUE (as printed)") or 0); check(fund == 0, f"{title} r{k}: on hold but priced")
            else:
                check(fund > 0, f"{title} r{k}: not on hold but fund price is 0")
            check(g("Proceeding") not in ("Falência", "RJ extinguished"), f"{title} r{k}: falência/extinct proceeding on a call tab")
            check(not re.search(r"MICRO|PEQUENO", str(g("Size") or ""), re.I) or "ME/EPP" in str(g("Flags") or "") or g("Proceeding", ).startswith("Recuperação extrajudicial"), f"{title} r{k}: ME/EPP creditor without the class IV flag")
            for ph in ("Company phone 1", "Company phone 2", "Backup phone", "Mobile (Apollo; reveal pending)"):
                check(not g(ph) or PHONE.match(str(g(ph))), f"{title} r{k}: phone not dialable as printed: {g(ph)!r}")
            check(not g("Backup phone") or g("Backup phone") not in (g("Company phone 1"), g("Company phone 2")), f"{title} r{k}: backup phone repeats the switchboard")
            check(g("STATUS") in STATUSES, f"{title} r{k}: status {g('STATUS')!r}")
            d = g("DEBTOR — party in RJ") or ""
            check(not re.search(r"administra[çc][ãa]o judicia|advocacia|trustee|p[áa]gina inicial|^ativos$|^base$|scalzilli", d, re.I), f"{title} r{k}: debtor looks like an administrator: {d[:50]}")
            check(not (g("PROOF QUALITY") or "").startswith("QUARANTINED"), f"{title} r{k}: raw status leaked into PROOF QUALITY")
            sh = g("Share of class III on this list")
            check(sh in (None, "") or (isinstance(sh, (int, float)) and 0 < sh <= 1.0001), f"{title} r{k}: share of class III {sh!r}")
            agio_sum += float(g("OUR ÁGIO (model, R$)") or 0)
        dvs = raw[title].data_validations.dataValidation
        check(any(dv.showErrorMessage for dv in dvs), f"{title}: data validation is advisory only")
        for row in raw[title].iter_rows(min_row=2, min_col=ix["PROOF: source document"] + 1, max_col=ix["PROOF: source document"] + 1):
            for c in row:
                if c.hyperlink and c.hyperlink.target: check(" " not in c.hyperlink.target, f"{title} {c.coordinate}: hyperlink target contains a space")
    d = wb["DASHBOARD"]
    kv = {d[f"A{i}"].value: d[f"B{i}"].value for i in range(1, 90) if d[f"A{i}"].value}
    check(kv.get("Leads in book (all tabs)") == total, f"dashboard leads {kv.get('Leads in book (all tabs)')} != rows {total}")
    check(kv.get("Rows on PRICING HOLD (zero quote until cleared)") == holds, f"dashboard holds {kv.get('Rows on PRICING HOLD (zero quote until cleared)')} != {holds}")
    check(abs(float(kv.get("Face on pricing hold (R$)") or 0) - face_hold) < 1, f"dashboard face on hold {kv.get('Face on pricing hold (R$)')} != {face_hold}")
    check(abs(float(kv.get("Model ÁGIO pipeline (R$)") or 0) - agio_sum) < 1, "dashboard ágio != sum of tabs")
    band_rows = {d[f"A{i}"].value: d[f"B{i}"].value for i in range(1, 90) if isinstance(d[f"A{i}"].value, str) and re.match(r"^[ABCDU] — ", d[f"A{i}"].value)}
    check(sum(band_rows.values()) == total, f"band block {sum(band_rows.values())} != rows {total}")
    top = [d[f"G{i}"].value for i in range(6, 21)]
    check(all(top[i] >= top[i + 1] for i in range(len(top) - 1) if isinstance(top[i], (int, float)) and isinstance(top[i + 1], (int, float))), "top 15 not descending")
    where = [d[f"K{i}"].value for i in range(6, 21)]
    check(len(set(w for w in where if w)) == len([w for w in where if w]), "top 15 shows the same row twice")
    rows_e = [(d[f"E{i}"].value, d[f"F{i}"].value) for i in range(22, 80)]
    all_i = next((i for i in range(22, 80) if d[f"E{i}"].value == "all rows"), None)
    check(all_i is not None and d[f"F{all_i}"].value == total, f"state block does not sum to the book ({all_i and d[f'F{all_i}'].value} vs {total})")
    check("_BOOK" in wb.sheetnames and wb["_BOOK"].sheet_state == "hidden", "_BOOK missing or visible")
    check("EXCLUDED (review)" in wb.sheetnames, "EXCLUDED tab missing")
    check(str(d["A2"].value or "").startswith("All call tabs counted together · run "), "build stamp missing")
shutil.rmtree(tmp, ignore_errors=True)
print(f"check_workbook: {len(fails)} failures" + ("" if not fails else "\n  " + "\n  ".join(fails[:25])))
sys.exit(1 if fails else 0)
