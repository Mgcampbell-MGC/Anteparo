"""The closer's workbook: DASHBOARD (live CRM formulas over every call tab) + CALL SHEET + FINANCIAL CREDITORS
+ STATE-OWNED CREDITORS + EXCLUDED (review) + Assumptions.

Ranking = model fund price (∝ our ágio), not face value. Every quote is an Excel formula off the Assumptions
tab keyed by DEBTOR BAND — or by the case's own plan terms when a plan PDF was read (plan_terms table) —
so re-pricing a band re-prices every lead in it. Proof-of-claim columns travel with the row.

Rules that protect the reader:
- the DEBTOR column never prints a registry name whose match was rejected; it prints the name two independent
  readers saw on the PDF head, else an accepted registry name, else a cleaned portal title, and says which;
- rows where the creditor is the debtor itself, a group member (by CNPJ, by a distinctive name token, or by a shared
  registered partner), rows filed under a subordinated / not-subject / class I-II-IV heading, rows below the printed
  class III total, ME/EPP creditors (class IV by art. 41 IV), falência or extinguished proceedings, and rows whose
  printed value could belong to a neighbouring creditor go to EXCLUDED (review), not the book;
- rows that cannot be priced honestly (non-statutory list, case not identified, printed name ≠ registry name for the
  printed CNPJ) stay in the book with a PRICING HOLD and a zero quote until the point is cleared;
- years-to-payment runs from the plan, not from the filing: an old case with no plan prices LOWER, not higher;
- band A means the recovery was GRANTED (art. 58); a plan that was merely filed (art. 53) is band B;
- PROOF QUALITY says whether the printed value was reconciled against the list's own totals.
"""
from __future__ import annotations

import json
import re
import subprocess
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from rapidfuzz import fuzz

STATUSES = ["NEW", "QUEUED", "EMAILED", "CALLED", "VOICEMAIL", "CONNECTED", "INTERESTED", "NOT INTERESTED", "CALL BACK", "DEAD",
            "TERMS AGREED", "CESSÃO SIGNED", "SUBSTITUTED IN LIST", "SETTLED"]
PIPELINE = ["INTERESTED", "TERMS AGREED", "CESSÃO SIGNED", "SUBSTITUTED IN LIST"]   # ágio counted as pipeline; SETTLED = realised
# band, class III recovery, grace (y), term (y), plan lag (y), meaning
#   plan lag: A = typical filing→grant delay used to estimate how far into the plan the case is;
#             B/C/U = years from TODAY until a plan is granted; D = years until a liquidation payout
BANDS = [("A", 0.40, 2, 8, 1.5, "Recovery GRANTED (art. 58): plan approved by the creditors and granted by the court; payments run from the grant"),
         ("B", 0.30, 2, 8, 1.5, "Live case filed within the last 2 years, plan filed (art. 53) or pending — not yet granted; the payment clock has not started"),
         ("C", 0.15, 3, 10, 2.5, "Older case (filed >2 y ago) or dormant/closed with no grant seen: a plan very probably exists but is not in our documents — check the docket before quoting"),
         ("D", 0.03, 0, 0, 6.0, "Not in the book: falência decreed, RJ extinguished, or the company is closed (baixada) at RFB — rows go to EXCLUDED (review)"),
         ("U", 0.25, 3, 8, 2.5, "Unresolved: debtor or case not identified — kept on PRICING HOLD (zero quote) until the case is found")]
BAND_SHORT = {"A": "A — recovery granted", "B": "B — live, plan pending", "C": "C — older, plan not seen", "D": "D — falência/closed (excluded)", "U": "U — unresolved (on hold)"}
BAND_MAP = {b: (rec, g, t, lag) for b, rec, g, t, lag, _ in BANDS}
DEFAULT_IRR, DEFAULT_SHARE, DEFAULT_FIN_MULT = 0.25, 0.60, 1.00
REJECTED_RFB = (None, "NAME_MISMATCH", "NOT_IN_RFB", "CREDITOR_CNPJ")
NAVY = "1F3A5F"; HEAD = PatternFill("solid", fgColor=NAVY); WHITE = Font(bold=True, color="FFFFFF")
FILL = {"crm": "FFF4D6", "deal": "E6EEF8", "proof": "E8F1E8", "contact": "F6F0FA", "ctx": "F3F3F3"}
BAND_FILL = {"A": "C6EFCE", "B": "E2EFDA", "C": "FFEB9C", "D": "FFC7CE", "U": "EDEDED"}
PROOF_FILL = {"NO TOTALS": "FFEB9C", "PARTIAL": "F8CBAD", "CHECK": "FF9999"}
TR = {"26": "SP", "19": "RJ", "16": "PR", "21": "RS", "24": "SC", "13": "MG", "09": "GO", "17": "PE", "05": "BA", "11": "MT", "12": "MS",
      "14": "PA", "15": "PB", "08": "ES", "10": "MA", "07": "DF", "06": "CE", "02": "AL", "25": "SE", "27": "TO", "20": "RN", "18": "PI",
      "22": "RO", "01": "AC", "04": "AM", "23": "RR", "03": "AP"}
LIST_KIND = {"AJ_LIST": "AJ list (art. 7 §2)", "DEBTOR_LIST": "Debtor list (art. 51)", "QGC": "QGC (art. 18) — homologation not verified", "PLAN": "Plan (art. 53) — not a list",
             "HOMOLOG": "Grant decision (art. 58)", "EDITAL": "Edital (art. 7 §2)", "UNKNOWN": "Unclassified list", "QUORUM_LIST": "Quorum list — NOT an art. 7/51 list",
             "OTHER": "Other filing — NOT an art. 7/51 list", "FALENCIA_LIST": "Falência creditor list (art. 99/83)"}
NOT_STATUTORY_KINDS = {"Quorum list — NOT an art. 7/51 list", "Other filing — NOT an art. 7/51 list", "Plan (art. 53) — not a list"}
DEBTOR_SOURCE = {"pdf-read": "printed on the list (read twice)", "rfb": "registry (RFB) match", "hint": "portal page title / URL — not printed on the list", None: "not identified"}
PROCEEDING = {"RJ": "Recuperação judicial", "RE": "Recuperação extrajudicial (art. 161)", "FALENCIA": "Falência", "EXTINCT": "RJ extinguished", "UNKNOWN": ""}
# CNAE 4-digit prefixes that are financial institutions (banks, co-ops, funds, factoring, securitisers, consórcio administrators, card/securities intermediaries)
FIN_CNAE = {"6410", "6421", "6422", "6423", "6424", "6431", "6432", "6433", "6434", "6435", "6436", "6437", "6438", "6440", "6450", "6461", "6470",
            "6491", "6492", "6493", "6499", "6611", "6612", "6613", "6630"}
STATE_NATUREZA = {"2011", "2038"}   # Empresa Pública, Sociedade de Economia Mista
FLAG_TEXT = {"PARTIAL_DOC_CLASS_III_RECONCILED": "class III reconciled (doc partial)", "UNRECONCILED_SOURCE": "list prints no totals",
             "CASE_NUMBER_CORRECTED": "case no. corrected by reader", "CASE_FROM_PORTAL": "case no. from portal record", "LIKELY_FINANCIAL_BY_NAME": "financial by name",
             "LIKELY_PUBLIC_BY_NAME": "public by name", "NO_RFB_RECORD": "no RFB record", "NO_ROUTE_IN": "no route in"}
DDD = {str(d) for d in (11, 12, 13, 14, 15, 16, 17, 18, 19, 21, 22, 24, 27, 28, 31, 32, 33, 34, 35, 37, 38, 41, 42, 43, 44, 45, 46, 47, 48, 49,
                        51, 53, 54, 55, 61, 62, 63, 64, 65, 66, 67, 68, 69, 71, 73, 74, 75, 77, 79, 81, 82, 83, 84, 85, 86, 87, 88, 89, 91, 92, 93, 94, 95, 96, 97, 98, 99)}
NONGEO = ("0800", "0300", "0500", "0900", "4004", "3003", "4003", "4020", "3004", "4007", "4090")

COLS = [  # (header, width, group)
    ("CREDITOR — company owed money", 40, "lead"), ("AMOUNT OWED (R$)", 17, "lead"), ("DEBTOR — party in RJ", 36, "lead"),
    ("STATUS", 18, "crm"), ("LAST TOUCH", 11, "crm"), ("NEXT ACTION", 22, "crm"), ("NEXT DATE", 11, "crm"), ("NOTES", 38, "crm"),
    ("DECISION MAKER", 30, "contact"), ("Role", 26, "contact"), ("Company phone 1", 18, "contact"), ("Company phone 2", 18, "contact"),
    ("Mobile (Apollo; reveal pending)", 16, "contact"), ("Email", 30, "contact"), ("Email status", 12, "contact"),
    ("BACKUP CONTACT", 28, "contact"), ("Backup role", 24, "contact"), ("Backup phone", 18, "contact"), ("Backup email", 28, "contact"),
    ("Contact note", 40, "contact"),
    ("FACE VALUE (as printed)", 17, "deal"), ("DEBTOR BAND", 8, "deal"), ("PRICING HOLD (why unpriced)", 34, "deal"),
    ("Plan: class III recovery", 10, "deal"), ("Plan: grace (y)", 8, "deal"), ("Plan: term (y)", 8, "deal"),
    ("Exp. recovery %", 9, "deal"), ("Years to payment", 8, "deal"),
    ("FUND PRICE @ IRR (model, R$)", 17, "deal"), ("Fund ¢/R$", 7, "deal"), ("OUR QUOTE (model, R$)", 17, "deal"), ("Quote ¢/R$", 7, "deal"),
    ("OUR ÁGIO (model, R$)", 16, "deal"), ("Ágio pts of face", 8, "deal"),
    ("PROOF QUALITY", 26, "proof"), ("PROOF: source document", 50, "proof"), ("Page / row (extractor index)", 16, "proof"), ("Value as printed", 30, "proof"),
    ("Creditor name as printed", 34, "proof"), ("Creditor CNPJ(s) as printed", 20, "proof"), ("Printed section heading", 30, "proof"), ("Recuperanda on the row (if printed)", 24, "proof"),
    ("Share of class III on this list", 9, "proof"), ("Class III total on list (R$)", 17, "proof"),
    ("Which list", 30, "proof"), ("List date", 11, "proof"), ("Plan evidence", 20, "proof"),
    ("Case number", 26, "ctx"), ("Court", 7, "ctx"), ("Proceeding", 18, "ctx"), ("Stage (DataJud)", 20, "ctx"), ("Filed", 7, "ctx"), ("Band reason", 44, "ctx"),
    ("Debtor CNPJ (lead entity)", 19, "ctx"), ("Debtor source", 30, "ctx"), ("Plan terms source", 30, "ctx"),
    ("Creditor type", 22, "ctx"), ("Seller status", 30, "ctx"),
    ("City", 18, "ctx"), ("UF", 5, "ctx"), ("Sector", 34, "ctx"), ("Size", 10, "ctx"), ("Capital (R$)", 15, "ctx"), ("Seller fit", 13, "ctx"),
    ("CNPJ root", 10, "ctx"), ("Claims", 6, "ctx"), ("Flags", 34, "ctx"),
]
C = {h: i + 1 for i, (h, _, _) in enumerate(COLS)}
L = lambda h: get_column_letter(C[h])


def _phone(p):
    """Registry phone → dialable text. Never invents a DDD: a number with a digit missing says so; 0800/4004 are printed as non-geographic."""
    d = "".join(ch for ch in str(p or "") if ch.isdigit())
    if not d or set(d) == {"0"}: return ""
    if d.startswith(("0800", "0300", "0500", "0900")):
        return f"{d[:4]} {d[4:7]} {d[7:]}" if len(d) == 11 else f"{d[:4]} {d[4:]}"
    if len(d) in (12, 13) and d.startswith("55"): d = d[2:]
    while d.startswith("0") and len(d) > 8: d = d[1:]     # trunk prefix ('0xx') or padding zeros
    if len(d) == 10 and d[:2] in DDD: return f"({d[:2]}) {d[2:6]}-{d[6:]}"
    if len(d) == 11 and d[:2] in DDD: return f"({d[:2]}) {d[2:7]}-{d[7:]}"
    if len(d) == 8: return f"{d[:4]}-{d[4:]}" + ("" if d.startswith(NONGEO) else " (no DDD)")
    if len(d) == 9:
        if d[0] == "9": return f"{d[:5]}-{d[5:]} (no DDD)"
        if d[:2] in DDD: return f"({d[:2]}) {d[2:6]}-{d[6:]} [digit missing]"
    return f"{d} [check]"


def _cnpj(d):
    d = "".join(ch for ch in str(d or "") if ch.isdigit())
    return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}" if len(d) == 14 else (d or "")


def _norm(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().upper()
    return re.sub(r"\s+", " ", re.sub(r"[^A-Z0-9 ]+", " ", s)).strip()


GENERIC = set("""LTDA LIMITADA SA EIRELI ME EPP EM RECUPERACAO JUDICIAL EXTRAJUDICIAL GRUPO DE DA DO DAS DOS CIA COMERCIO COMERCIAL INDUSTRIA INDUSTRIAL IND COM
SERVICOS SERVICO PARTICIPACOES PARTICIPACAO EMPREENDIMENTOS EMPREENDIMENTO BRASIL BRASILEIRA BRASILEIRO ADMINISTRACAO ADMINISTRADORA HOLDING OUTRAS OUTROS
MASSA FALIDA FALIDO PRODUTOS PRODUTO AGRICOLA AGRICOLAS AGRO AGROPECUARIA AGROINDUSTRIAL AGRONEGOCIO ALIMENTOS ALIMENTICIOS TRANSPORTES TRANSPORTE LOGISTICA
TEXTIL CONFECCOES CONFECCAO MALHAS MALHARIA DISTRIBUIDORA DISTRIBUICAO IMPORTACAO EXPORTACAO IMPORTADORA EXPORTADORA EQUIPAMENTOS MAQUINAS EMBALAGENS PLASTICOS
METALURGICA METAIS CONSTRUCOES CONSTRUTORA CONSTRUCAO ENGENHARIA INCORPORADORA IMOBILIARIA EMPRESA EMPRESAS NACIONAL INTERNACIONAL GERAL GERAIS CENTRAL CENTRO
NORTE SUL LESTE OESTE NOVA NOVO SAO SANTA SANTO BANCO COOPERATIVA CREDITO FUNDO INVESTIMENTO INVESTIMENTOS FIDC SECURITIZADORA FACTORING FOMENTO MERCANTIL
CONSULTORIA SOLUCOES TECNOLOGIA SISTEMAS ENERGIA PETROLEO GAS COMBUSTIVEIS QUIMICA FARMACEUTICA MEDICAMENTOS HOSPITALARES SAUDE MOVEIS MADEIRAS PAPEL CELULOSE
CALCADOS COUROS BEBIDAS FRIGORIFICO PESCADOS RACOES RACAO FABRICA PECAS AUTO AUTOMOTIVA VEICULOS TERCEIRIZACAO VIGILANCIA SEGURANCA LIMPEZA GESTAO ATACADO
ATACADISTA VAREJO SUPERMERCADOS SUPERMERCADO MATERIAIS MATERIAL FERRAGENS ACO FERRO ELETRICA ELETRICOS ELETRONICOS TELECOMUNICACOES REPRESENTACOES REPRESENTACAO
ASSESSORIA CONTABILIDADE ADVOGADOS SOCIEDADE ANONIMA EDITAL CNPJ CPF PRODUTOR RURAL FILIAL MATRIZ UNIDADE COMPANHIA DIREITOS CREDITORIOS MULTISETORIAL
INDUSTRIAS AGRICULTURA PECUARIA SEMENTES FERTILIZANTES DEFENSIVOS INSUMOS NUTRICAO ANIMAL AVICOLA AVICULTURA SUINOS LATICINIOS USINA ACUCAR ALCOOL ETANOL
POSTO POSTOS DERIVADOS LUBRIFICANTES PNEUS AUTOPECAS TINTAS VIDROS CERAMICA CIMENTO ARTEFATOS TECIDOS FIOS FIACAO TECELAGEM ESTAMPARIA CALCADISTA BOLSAS
JOIAS ELETRODOMESTICOS INFORMATICA SOFTWARE TELECOM INTERNET PROVEDOR SEGUROS SEGURADORA CORRETORA CAPITAL PREVIDENCIA CONSORCIO ADMINISTRADORES LOCACAO
LOCADORA ALUGUEL HOTEL HOTEIS TURISMO VIAGENS EDUCACAO ENSINO ESCOLA COLEGIO FACULDADE UNIVERSIDADE CLINICA HOSPITAL LABORATORIO DIAGNOSTICOS ODONTO
MEDICA MEDICOS FARMACIA DROGARIA COSMETICOS PERFUMARIA HIGIENE LIMPEZA SANEAMENTO AMBIENTAL RESIDUOS RECICLAGEM SUCATA MINERACAO MINERIOS PEDRAS
GRANITOS MARMORES MADEIREIRA SERRARIA FLORESTAL CELULOSE GRAFICA EDITORA PUBLICIDADE PROPAGANDA EVENTOS PROMOCOES PROMOCOES MARKETING COMUNICACAO
SHOPPING CENTER EMPREENDIMENTOS IMOVEIS PARTICIPACOES PATRIMONIAL FAMILIA IRMAOS FILHOS FILHO NETO JUNIOR SOBRINHO
RESPONSABILIDADE NAO PADRONIZADOS PADRONIZADO MULTISSETORIAL MULTISETORIAL SENIOR MEZANINO COTAS CLASSE CREDITORIOS UNICA FUNDOS ABERTO FECHADO
COOPERATIVO COOPERATIVAS LIVRE ADMISSAO ASSOCIADOS POUPANCA RURAL EMPRESARIOS MICROEMPRESARIOS PROFISSIONAIS MEDICOS DEMAIS SAUDE PROF EMP CONT""".split())


GEO = set("""ACRE ALAGOAS AMAPA AMAZONAS BAHIA CEARA ESPIRITO GOIAS MARANHAO MATO GROSSO MINAS GERAIS PARA PARAIBA PARANA PERNAMBUCO PIAUI RIO JANEIRO GRANDE NORTE SUL
RONDONIA RORAIMA SANTA CATARINA SERGIPE TOCANTINS DISTRITO FEDERAL BRASILIA PAULISTA PAULISTANO CARIOCA FLUMINENSE GAUCHO GAUCHA CATARINENSE PARANAENSE MINEIRO MINEIRA
BAIANO BAIANA CAPIXABA GOIANO GOIANA MATOGROSSENSE PERNAMBUCANO CEARENSE POTIGUAR PARAIBANO ALAGOANO SERGIPANO MARANHENSE PIAUIENSE PARAENSE AMAZONENSE TOCANTINENSE
NORDESTE NORDESTINO CENTRO OESTE SUDESTE ALEGRE PORTO VALE ALTO BAIXO CAMPO CAMPOS SERRA LAGOA MONTE MONTES VILA PARK PARQUE CENTER SHOPPING CLUB CLUBE CONDOMINIO
CONSORCIO UNIMED SICOOB SICREDI UNICRED CRESOL CENTRAL PLANALTO LITORAL FRONTEIRA MISSOES PAMPA CERRADO SERTAO AGRESTE TRIANGULO URUGUAI PARAGUAI ARGENTINA
BLUMENAU JOINVILLE CURITIBA LONDRINA MARINGA CASCAVEL CHAPECO CRICIUMA CAXIAS PELOTAS PASSO FUNDO CAMPINAS RIBEIRAO PRETO SOROCABA SANTOS BAURU PIRACICABA UBERLANDIA
UBERABA JUIZ FORA CONTAGEM BETIM GOIANIA ANAPOLIS CUIABA RONDONOPOLIS DOURADOS RECIFE CARUARU PETROLINA FORTALEZA SALVADOR FEIRA SANTANA VITORIA LINHARES BELEM MANAUS
NATAL JOAO PESSOA MACEIO ARACAJU TERESINA LUIS PALMAS VELHO BRANCO BOA VISTA MACAPA ITAJAI FLORIANOPOLIS LAGES JARAGUA BRUSQUE ERECHIM CONCORDIA JOACABA
CIDADE ESTADO MUNICIPIO PREFEITURA REGIAO REGIONAL NACIONAL UNIAO UNIDOS UNIVERSAL GLOBAL MUNDIAL MODERNA MASTER PLUS PRIME REAL CRISTAL BEM BOM MAX TOP
GRAOS CANA PLANTADORES PRODUTORES AGRICULTORES CONDUTORES ELETRICOS CABOS FIOS""".split())
_NOT_NAME = re.compile(r"^(SIM|NAO|N A|PENDENTE.*|TITULARES.*|CLASSE .*|QUIROGRAF.*|CREDORES.*|VALOR.*|TOTAL.*)$|"
                       r"\b(AV|AVENIDA|RUA|ROD|RODOVIA|ALAMEDA|AL|PRACA|PCA|PC|QUADRA|SETOR|KM|CEP|ANDAR|SALA|CONJ|CJ|LOTE|BLOCO|BL|ESTRADA|TRAVESSA|NUC|BAIRRO|CENTRO|N)\b")


def _tokens(s):
    return [t for t in _norm(s).split() if (len(t) >= 3 or (len(t) == 2 and any(ch.isdigit() for ch in t))) and not t.isdigit()]


def _is_name(printed):
    """True when the extractor's 'creditor as printed' is a company name and not an address, a SIM/NÃO column, a legend or a number."""
    p = _norm(printed)
    return bool(p) and not _NOT_NAME.search(p) and len(re.sub(r"[^A-Z]", "", p)) >= 4 and len(p) <= 90


def name_mismatch(printed, razao, generic, fantasia=""):
    """'' when the printed creditor name and the registry name (razão or nome fantasia) for the printed CNPJ agree — abbreviations,
    initials and reorderings tolerated. A printed name made only of generic words ('BRASIL S.A') cannot vouch for a specific registry name."""
    if not razao or not _is_name(printed): return ""
    pt = _tokens(printed); pt_dist = [t for t in pt if t not in generic]
    ns = lambda x: re.sub(r"\s+", "", _norm(x))
    for reg in (razao, fantasia or ""):
        if not reg: continue
        rt = [t for t in _tokens(reg) if t not in generic]
        if not rt: return ""
        if any(t in pt or any(len(q) >= 3 and (t.startswith(q) or q.startswith(t)) for q in pt) for t in rt): return ""
        if pt_dist and fuzz.partial_ratio(ns(printed), ns(reg)) >= 80: return ""
        initials = "".join(w[0] for w in _norm(printed).split() if len(w) == 1)
        if initials and initials in rt: return ""
    return f"list prints '{str(printed).strip()[:40]}' but the printed CNPJ is registered to {razao[:45]}" + (f" (fantasia {fantasia[:30]})" if fantasia else "")


def _seller_fit(razao, capital, porte):
    """LARGE CORP = the Petrobras/Hypera tier that sells through a desk or not at all; UNKNOWN when RFB has nothing."""
    if not razao: return "UNKNOWN"
    try: cap = float(capital or 0)
    except ValueError: cap = 0
    if not capital: return "UNKNOWN"
    if cap >= 1_000_000_000 or (re.search(r"\bS\.?A\.?\s*$|\bS/A\s*$|\bS\.A\.\b", razao) and cap >= 300_000_000):
        return "LARGE CORP"
    if porte and ("MICRO" in porte.upper() or "PEQUENO" in porte.upper()):
        return "SMALL"
    return "MID-MARKET"


def _seller_status(razao, name_printed, situacao, situacao_desc):
    """What stands between a yes and a signature on the SELLER's side."""
    blob = f"{razao or ''} {name_printed or ''}"
    if re.search(r"MASSA FALIDA|\bFALID[AO]\b", blob, re.I): return "FALIDA — only its AJ can sell (court authorisation)"
    if re.search(r"LIQUIDA[CÇ][AÃ]O", blob, re.I): return "EM LIQUIDAÇÃO — only the liquidante (BCB/court) can sell"
    if re.search(r"RECUPERA[CÇ][AÃ]O (JUDICIAL|EXTRAJUDICIAL)", blob, re.I): return "IN RJ ITSELF — its own judge/AJ must authorise the sale (art. 66)"
    if not razao: return "no RFB record — status unknown"
    if situacao and situacao != "02": return f"RFB {situacao_desc or situacao} — confirm who signs"
    return "ATIVA"


_ADMIN = re.compile(r"administra[çc][ãa]o judicia|administradora judicia|administrador judicia|administra[çc][õo]es judicia|\bAJ\b|advogad|advocacia|"
                    r"sociedade de advogados|consultoria|escrit[óo]rio|per[íi]cia|l[íi]deres em recupera|recupera[çc][õo]es judiciais|p[áa]gina inicial|"
                    r"\bbase\s*\||^ativos\b|^base$|trustee|gatekeeper|credibilit|scalzilli|sgrott|diligence|lindoso|stenius|\bruiz\b|cat[áa]lise|recupera solu|"
                    r"biancardi|vivante|inova-?aj|rlg-?aj|\bvtl\b|assistjud|marques adm|registrado civilmente", re.I)
_PREFIX = re.compile(r"^\s*(?:recupera[çc][ãa]o judicial (?:de|da|do|das|dos)?|rj (?:de|da|do)?|massa falida (?:de|da|do)?|fal[êe]ncia (?:de|da|do)?|"
                     r"processo(?: n[º°.]*)?\s*[\d.\-/]+\s*[-–—:]?|e fal[êe]ncia|aludida empresa|empresa|promovida por)\s*", re.I)
_SUFFIX_RJ = re.compile(r"\s*[-–—(]?\s*em recupera[çc][ãa]o judicial\)?\s*$", re.I)
_CASE = re.compile(r"\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}")
_SPLIT = re.compile(r"\s+[–—|]\s+|\s+-\s+")
_ATYPICAL = re.compile(r"\bASSOCIA[ÇC][ÃA]O\b|\bCLUBE\b|FOOTBALL|FUTEBOL|SPORT CLUB|SANTA CASA|HOSPITAL|FUNDA[ÇC][ÃA]O|PRODUTOR RURAL|PESSOA F[ÍI]SICA|\bCPF\b", re.I)
_ROOT_IN_TEXT = re.compile(r"(\d{2})\.?(\d{3})\.?(\d{3})/?\d{4}-?\d{2}")


def clean_hint(h):
    """Portal page title / debtor hint → a company name, or '' when it is not one (administrator's site name, case number, prose)."""
    if not h: return ""
    h = re.sub(r"\s+", " ", str(h)).strip()
    keep = [p for p in _SPLIT.split(h) if p and not _ADMIN.search(p) and not _CASE.fullmatch(p.strip())]
    if not keep: return ""
    h = _SUFFIX_RJ.sub("", _PREFIX.sub("", keep[0])).strip(" -–—:|,")
    if len(h) < 4 or _ADMIN.search(h) or (_CASE.search(h) and len(h) < 40): return ""
    prose = len(re.findall(r"\b(litiscons[óo]rcio|conforme|entendimento|requerimento|foi|apresentad[oa]|autos|conson[âa]ncia|movimento|supramencionad[oa]s?|"
                           r"referid[oa]|aludid[oa]|mediante|sendo|seja|cuj[oa]|qual|pelo|pela|nos|nas|para que|chegue|civilmente|r[ée]u)\b", h, re.I))
    if prose >= 1 and not re.search(r"\b(LTDA|S\.?A\.?|S/A|EIRELI|EPP|ME|GRUPO)\b", h, re.I): return ""
    if prose >= 2: return ""
    if not re.search(r"\b[A-ZÀ-Ý][A-ZÀ-Ýa-zà-ÿ&'.]{2,}", h): return ""
    return h


def debtor_display(display, razao, src, *hints):
    """What the closer sees as the debtor. Never a registry name whose match was rejected."""
    if display: return display
    if razao and src not in REJECTED_RFB and not _ADMIN.search(razao): return razao
    if razao and src == "NAME_MISMATCH" and re.search(r"RECUPERA[CÇ][AÃ]O JUDICIAL", razao, re.I) and not _ADMIN.search(razao):
        return razao   # the registry itself carries the RJ suffix: it is the recuperanda even though the page title did not match
    for h in hints:
        c = clean_hint(h)
        if c: return c
    return ""


def _years(band, ysf, plan=None):
    rec, g, term, lag = BAND_MAP.get(band, BAND_MAP["U"])
    if plan: g, term = plan[1], plan[2]
    if band == "A":
        return max(1.0, g + term / 2 - max(0.0, ysf - lag))
    return lag + g + term / 2


def _related(root, cred_names, debtor_root, member_roots, debtor_names, cred_partners, debtor_partners, generic):
    """Why the creditor is (probably) the debtor's own family. '' when nothing links them."""
    if debtor_root and root == debtor_root: return "creditor has the debtor's own CNPJ (intercompany)"
    if root in member_roots: return "creditor CNPJ is listed among the recuperandas (group member)"
    cred_names = [n for n in cred_names if n and (_is_name(n) or n == cred_names[0])]
    dist = lambda n: [t for t in _tokens(n) if t not in generic and t not in GEO]
    heads = {d[0] for d in (dist(n) for n in cred_names) if d}          # the first distinctive word of the creditor's name
    cred_tok = {t for n in cred_names for t in dist(n)}
    for nm in debtor_names:
        if not nm: continue
        dt = dist(nm)
        shared = [t for t in dt if t in cred_tok and (t in heads or t == dt[0])]   # must be the name's head word on at least one side
        if shared:
            return f"shares the distinctive name '{shared[0]}' with the debtor / a group member ({nm[:40]})"
        for cand in cred_names:
            if cand and len(dt) >= 2 and fuzz.token_set_ratio(_norm(nm), _norm(cand)) >= 90:
                return f"creditor name = debtor / group member ({nm[:40]})"
    common = cred_partners & debtor_partners
    if common:
        return f"shares a registered partner with the debtor ({sorted(common)[0].title()})"
    return ""


TIER_A = re.compile(r"financ|\bcfo\b|controller|controlad|tesour|treasur|cr[eé]dito|credit|cobran|collect|receiv|contas a receber|recupera[cç][aã]o de|special assets|"
                    r"workout|\bnpl\b|reestrutura|restructur|s[oó]ci[oa]|partner|owner|propriet|\bdono\b|fundador|founder|presidente|president|\bceo\b|diretor|director|"
                    r"gerente geral|general manager|administrador|superintend|head of f|\bcoo\b|managing", re.I)
TIER_NOT = re.compile(r"marketing|recursos humanos|\brh\b|human resources|seguran[çc]a|safety|\bti\b|tecnologia|information technology|vendas|sales|comercial|commercial|"
                      r"log[íi]stic|dep[óo]sito|warehouse|farmac|assistente|assistant|estagi|intern\b|analista de sistemas|engenheir|técnic|tecnic|supervisor de (produ|merch)|"
                      r"customer|design|qualidade", re.I)


def _tier(title):
    t = title or ""
    if TIER_A.search(t) and not (TIER_NOT.search(t) and not re.search(r"financ|cr[eé]dito|cobran|controller|tesour", t, re.I)): return "A"
    if re.search(r"gerente|manager|coordenador|coordinator", t, re.I) and not TIER_NOT.search(t): return "B"
    return "C"


# --- section headings ---------------------------------------------------------------------------------------------
_H_LEGEND = re.compile(r"TRABALHIST|QUIROGRAF|GARANTIA REAL|EXTRACONCURS|CONCURSAL|CLASSE I\b|CLASSE II\b|CLASSE III\b|CLASSE IV\b")
_H_SUB = re.compile(r"SUBORDINAD")
_H_OUT = re.compile(r"NAO SUJEIT|EXTRACONCURS|CREDITOS? FISCA|TRIBUTAR|OBRIGAC(?:OES|AO) DE (?:FAZER|ENTREGAR)|\bENTREGAR\b|CLASSE VI\b|LISTA VI\b|\bILIQUID|SUJEITO A CONDIC|RETIDO")
_H_OTHER = re.compile(r"CLASSE I\b(?! ?I)|CLASSE I [-–—:]|TRABALHIST|CLASSE II\b(?! ?I)|GARANTIA REAL|\bLEASING\b|CLASSE IV\b|ME E EPP|ME ?/ ?EPP|MICROEMPRESA|PEQUENO PORTE|"
                      r"(?:SUB)?TOTAL (?:DA |DE |GERAL DA )?CLASSE III|TOTAL CLASSE III")


def heading_verdict(heading, class3_reconciled, set_by="SECTION_HEADING"):
    """'' when the row may stay; otherwise why a row under this printed heading is not tradeable class III paper.
    set_by = how the extractor fixed the row's class: SECTION_HEADING (inherited from this heading), COLUMN / INLINE_HEADER (printed on the row itself,
    so a contradicting heading is a carry-over and is ignored)."""
    h = _norm(heading)
    if not h: return ""
    if len(set(_H_LEGEND.findall(h))) >= 3: return ""          # a column legend naming every class, not a section heading
    rowlike = bool(re.search(r"\d{1,3}(?:\.\d{3})+,\d{2}", heading or "")) and not re.search(r"TOTAL|CREDOR", h)   # a creditor row carried as heading
    if rowlike or len(h) > 160: return ""
    if _H_SUB.search(h) and "PRIVILEGIO" not in h: return "listed under a SUBORDINADO heading (art. 83 VIII: partners' and insiders' claims, paid last) — not tradeable class III"
    if _H_OUT.search(h): return f"listed under a heading that is not concursal class III ({heading.strip()[:60]})"
    if set_by == "SECTION_HEADING" and _H_OTHER.search(h):
        # the class was inherited from this heading; when the heading is the class III TOTAL line the row sits in the next section,
        # whatever any per-section sum says (a section total can match the next class's own total)
        if re.search(r"TOTAL", h): return f"row printed below the class III total line ({heading.strip()[:50]}) — belongs to the next class"
        if not class3_reconciled: return f"listed under a class I / II / IV heading on the source list ({heading.strip()[:50]}) — not class III"
    return ""


def list_kind(dtype, url, title, link):
    blob = _norm(f"{url or ''} {title or ''} {link or ''}")
    if re.search(r"QUORUM", blob): return LIST_KIND["QUORUM_LIST"]
    if dtype == "PLAN" or (re.search(r"\bPLANO\b|\bPRJ\b", blob) and not re.search(r"RELAC|LISTA|QUADRO|EDITAL|CREDORES", blob)): return LIST_KIND["PLAN"]
    if dtype == "OTHER": return LIST_KIND["OTHER"]
    if re.search(r"HOMOLOG|CONCESS", blob) and not re.search(r"QGC|QUADRO", blob): return LIST_KIND["HOMOLOG"]
    if re.search(r"PROVIS", blob) and re.search(r"QGC|QUADRO", blob): return "Provisional QGC (AJ consolidation, not homologated)"
    if re.search(r"HOMOLOG", blob) and re.search(r"QGC|QUADRO", blob): return "QGC homologated (art. 18)"
    if re.search(r"\bART(?:IGO)? ?7\b|7 ?[O°º] ?(?:PARAGRAFO|§)|2 ?[A°º] ?(?:LISTA|RELACAO)|SEGUNDA (?:LISTA|RELACAO)|LISTA DO ADMINISTRADOR|RELACAO DO ADMINISTRADOR|\bAJ\b.*(?:LISTA|RELACAO)", blob): return LIST_KIND["AJ_LIST"]
    if re.search(r"\bART(?:IGO)? ?51\b|RELACAO DE CREDORES DA RECUPERANDA|APRESENTADA PELA RECUPERANDA|LISTA DA RECUPERANDA|RELACAO DE CREDORES DA DEVEDORA", blob): return LIST_KIND["DEBTOR_LIST"]
    if re.search(r"MASSA FALIDA|\bFALIDA\b|\bART(?:IGO)? ?99\b|CREDORES DA FALENCIA", blob): return LIST_KIND["FALENCIA_LIST"]
    return LIST_KIND.get(dtype or "UNKNOWN", dtype or "")


def _recon(rj):
    """Class III reconciliation facts from documents.reconcile_json: (reconciled?, printed class III total or None, delta text)."""
    try: j = json.loads(rj or "{}")
    except Exception: j = {}
    chs = [c for c in j.get("checks", []) if str(c.get("class", "")).startswith("III/BRL")]   # one check per printed section ('III/BRL#s3')
    if not chs: return None, None, ""
    tot = Decimal(0)
    for c in chs:
        try: tot += Decimal(str(c.get("printed_total")))
        except Exception: pass
    deltas = [str(c.get("delta")) for c in chs if c.get("delta") not in (None, "0.00")]
    return all(c.get("pass") for c in chs), (tot if tot > 0 else None), (f"delta R$ {'; '.join(deltas)}" if deltas else "")


def _partners(qsa_json):
    try: q = json.loads(qsa_json or "[]")
    except Exception: q = []
    out = set()
    for p in q:
        n = _norm(p.get("nome") or p.get("nome_socio") or "")
        if n and len(n.split()) >= 2 and not re.search(r"LTDA|\bSA\b|EIRELI|PARTICIPACOES|HOLDING", n): out.add(n)
    return out


def _leads(db, run_id):
    plan_terms = {r[0]: (r[1], r[2], r[3], r[4]) for r in db.execute("SELECT case_number, recovery, grace, term, source FROM plan_terms")} if \
        db.execute("SELECT name FROM sqlite_master WHERE name='plan_terms'").fetchone() else {}
    case_notes = {r[0]: r[1] for r in db.execute("SELECT case_number, note FROM case_notes")} if db.execute("SELECT name FROM sqlite_master WHERE name='case_notes'").fetchone() else {}
    # tokens that name an industry rather than a company: frequent across the creditor corpus, so never evidence of kinship
    df = Counter()
    for (rz,) in db.execute("SELECT razao_social FROM companies WHERE razao_social IS NOT NULL"):
        df.update(set(_tokens(rz)))
    generic = GENERIC | {t for t, n in df.items() if n >= 6}
    rfb_qsa = {}
    def debtor_qsa(cnpj):
        if not cnpj: return set()
        if cnpj not in rfb_qsa:
            row = db.execute("SELECT json FROM rfb_cache WHERE cnpj=? AND json IS NOT NULL", (cnpj,)).fetchone()
            try: rfb_qsa[cnpj] = _partners(json.dumps(json.loads(row[0]).get("qsa") or [])) if row else set()
            except Exception: rfb_qsa[cnpj] = set()
        return rfb_qsa[cnpj]
    q = db.execute("""
        SELECT t.cnpj_basico, t.case_number, t.doc_id, t.class_iii_face_sum, t.establishment_cnpjs, t.claim_count, t.creditor_name_as_printed, t.flags,
               c.razao_social, c.phone, c.phone2, c.email, c.uf, c.municipio, c.cnae_desc, c.porte, c.capital_social, c.is_bank, c.is_public, c.is_inactive,
               c.cnae_principal, c.natureza_juridica, c.nome_fantasia, c.qsa_json, c.situacao_cadastral, c.situacao_desc,
               d.source_url, d.doc_type, d.publication_date, d.status, d.file_path, d.debtor_name_hint, d.notes, d.reconcile_json,
               r.page_title, r.link_text,
               cs.court, cs.stage, cs.filing_date, cs.rj_granted_signal,
               db2.debtor_name, db2.razao_social, db2.band, db2.band_reasons, db2.plan_status, db2.filing_year, db2.stage, db2.rfb_source, db2.debtor_cnpj,
               db2.display_name, db2.verified_cnpj, db2.group_members, NULLIF(dd.debtor,''), db2.display_source, db2.proceeding
        FROM targets t
        LEFT JOIN companies c ON c.cnpj_basico=t.cnpj_basico
        LEFT JOIN documents d ON d.doc_id=t.doc_id AND d.run_id=t.run_id
        LEFT JOIN raw_documents r ON r.url=d.source_url
        LEFT JOIN cases cs ON cs.case_number=t.case_number
        LEFT JOIN debtors db2 ON db2.case_number=t.case_number
        LEFT JOIN doc_debtor dd ON dd.doc_id=t.doc_id
        WHERE t.run_id=? AND t.band='FLOOR'""", (run_id,)).fetchall()
    out, excluded = [], []
    for r in q:
        (root, case, doc_id, face, ests, ncl, name_printed, flags, razao, phone, phone2, email, uf, city, cnae, porte, capital,
         is_bank, is_public, is_inactive, cnae_code, natureza, fantasia, qsa_json, situacao, situacao_desc,
         url, dtype, pub, dstatus, fpath, dhint, dnotes, recon_json, ptitle, plink,
         court, stage, filed, granted,
         dname, drazao, band, breason, pstatus, fyear, dstage, dsrc, dcnpj, ddisplay, vcnpj, gmembers, dd_debtor, dsource, proceeding) = r
        flags = flags or ""; face_d = Decimal(face)
        ests_l = (ests or "").split("|")
        debtor = debtor_display(ddisplay, drazao, dsrc, dd_debtor, dname, dhint)
        fin = ((cnae_code or "")[:4] in FIN_CNAE) if cnae_code else (bool(is_bank) or "LIKELY_FINANCIAL" in flags)
        state = (natureza or "") in STATE_NATUREZA or (not razao and "LIKELY_PUBLIC" in flags)   # a state bank with no RFB record still sells by desk/auction
        book = "STATE-OWNED" if state else ("FINANCIAL" if fin else "TRADE")
        ctype = ("STATE-OWNED (financial)" if fin else "STATE-OWNED") if state else ("FINANCIAL (bank/FIDC/co-op)" if fin else "TRADE")
        base = {"razao": razao, "name_printed": name_printed, "face": face_d, "case": case or "", "url": url, "debtor": debtor, "root": root, "book": book}
        # --- exclusions (visible on the EXCLUDED tab) ---
        public = (natureza or "").startswith("1") or (bool(is_public) and (natureza or "").startswith("1"))
        name_public = not razao and "LIKELY_PUBLIC" in flags and not re.search(r"\bBANCO\b|\bCAIXA\b|\bBANK\b", name_printed or "", re.I)
        if public or name_public:
            excluded.append({**base, "reason": "public body (administração pública) — not a seller" if public else "public by name, no RFB record — verify before calling"}); continue
        if is_inactive:
            excluded.append({**base, "reason": f"creditor company not ATIVA at RFB ({situacao_desc or situacao or 'status unknown'})"}); continue
        if porte and re.search(r"MICRO|PEQUENO", porte.upper()):
            excluded.append({**base, "reason": f"creditor is {porte.title()} at RFB — an ME/EPP creditor votes in class IV (art. 41 IV), so this is class IV paper even if the list prints it under class III; verify class before quoting"}); continue
        accepted_cnpj = dcnpj if dsrc not in REJECTED_RFB else ""
        debtor_root = (vcnpj or accepted_cnpj or "")[:8]
        members = json.loads(gmembers) if gmembers else []
        member_roots = {"".join(m.groups()) for mm in members for m in _ROOT_IN_TEXT.finditer(mm or "")}
        cred_partners = _partners(qsa_json)
        d_partners = debtor_qsa(vcnpj or accepted_cnpj) | {_norm(re.sub(r"\(.*?\)", "", m)) for m in members if re.search(r"\bCPF\b|PRODUTOR RURAL|PESSOA F", m or "", re.I)}
        rel = _related(root, [razao, name_printed], debtor_root, member_roots, [debtor, drazao if dsrc not in REJECTED_RFB else ""] + members, cred_partners, d_partners, generic)
        cred_in_rj = bool(re.search(r"RECUPERA[CÇ][AÃ]O JUDICIAL", razao or "", re.I))
        if rel:
            excluded.append({**base, "reason": f"related party: {rel}"}); continue
        claims = db.execute("""SELECT page, row_index, value_as_printed, value_brl, section_heading, flags, all_documents, debtor_as_printed, class_set_by FROM claims
                               WHERE doc_id=? AND run_id=? AND class='III' AND currency='BRL' AND superseded_by IS NULL AND document_number IN (%s) ORDER BY page,row_index"""
                            % ",".join("?" * len(ests_l)), (doc_id, run_id, *ests_l)).fetchall()
        iii_pass, iii_total, iii_delta = _recon(recon_json)
        class3_reconciled = (dstatus == "OK") or bool(iii_pass)
        verdict = next((v for v in (heading_verdict(c[4], class3_reconciled, c[8]) for c in claims) if v), "")
        if verdict:
            excluded.append({**base, "reason": verdict}); continue
        other_roots = {x[:8] for c in claims for x in (c[6] or "").split("|") if x and x[:8] != root}
        if any("MULTI_ROOT" in (c[5] or "") for c in claims) or other_roots:
            excluded.append({**base, "reason": "printed value sits in a prose line naming another CNPJ — may belong to a neighbouring creditor; re-extract before quoting"}); continue
        # --- classification / banding ---
        if not case or case[16:17] != "8":   # RJ cases live in state courts (J=8); anything else is a creditor-side number picked off the page
            case = ""; band = "U"; breason = "RJ case number not identified on the document"; stage = stage or ""
        band = band if band in BAND_MAP else "U"
        proc = proceeding or ""
        if band == "D" or proc in ("FALENCIA", "EXTINCT"):
            why = {"FALENCIA": "falência: claims rank under art. 83 and are paid from the liquidation — not RJ class III paper",
                   "EXTINCT": "the RJ was extinguished — there is no plan to buy into"}.get(proc, "debtor closed (baixada) at RFB / falência decreed")
            excluded.append({**base, "reason": f"{why}; {breason or ''}".rstrip("; ")}); continue
        if proc == "RE": breason = ((breason + "; ") if breason else "") + "recuperação extrajudicial (art. 161): plan is a private agreement"
        if _ATYPICAL.search(debtor or ""): breason = ((breason + "; ") if breason else "") + "atypical debtor (associação/clube/hospital/produtor rural): admissibility or claim subjection may be contested"
        if "CASE_FROM_PORTAL" in flags or "CASE_FROM_PAGE" in (dnotes or ""): breason = ((breason + "; ") if breason else "") + "case no. from the portal record, not printed on the list"
        if case and case_notes.get(case): breason = ((breason + "; ") if breason else "") + case_notes[case]
        year = fyear or (int(filed[:4]) if filed else None)
        if not year:
            m = re.match(r"\d{7}-\d{2}\.(\d{4})\.", case or ""); year = int(m.group(1)) if m else None
        ysf = (date.today().year - year) if year else 0
        plan = plan_terms.get(case) if case else None
        # --- proof ---
        kind = list_kind(dtype, url, ptitle, plink)
        listed = sum(float(c[3]) for c in claims if c[3] is not None)
        if dstatus == "OK": proof = "OK — totals reconciled"
        elif dstatus == "OK_NO_TOTALS": proof = "NO TOTALS — list prints none; check page"
        elif dstatus == "QUARANTINED": proof = "PARTIAL — class III reconciled, other classes not" if iii_pass else f"CHECK — class III total did not reconcile ({iii_delta or 'no class III total found'})"
        else: proof = dstatus or ""
        if kind in NOT_STATUTORY_KINDS: proof = "CHECK — source is not a statutory creditor list"
        vals = []
        for pg, ri, v, vb, hd, cf, ad, dp, sb in claims:
            if v is None: continue
            alt = re.search(r"ALT_VALUE:([^|]+)", cf or "")
            vals.append(f"p{pg}: R$ {v}" + (f" (also printed: R$ {alt.group(1).strip()})" if alt else ""))
        vals_printed = "; ".join(vals)
        two_values = any("ALT_VALUE" in (c[5] or "") for c in claims)
        if claims and abs(listed - float(face_d)) > 0.01 * float(face_d): vals_printed = f"[rows shown ≠ face] {vals_printed}"
        headings = [c[4].strip() for c in claims if c[4]]
        recup_rows = sorted({(c[7] or "").strip() for c in claims if c[7]})
        mismatch = name_mismatch(name_printed, razao, generic, fantasia)
        share = (float(face_d) / float(iii_total)) if iii_total else None
        if share is not None and share > 1.001: share = None; extra_share_note = "face exceeds the printed class III total — the printed total covers a section, not the class"
        else: extra_share_note = ""
        hold = []
        if kind in NOT_STATUTORY_KINDS: hold.append("source is not a statutory creditor list — find the art. 7 §2 / art. 51 list first")
        if band == "U": hold.append("RJ case not identified — cannot place the claim in a proceeding")
        if not debtor and not case: hold.append("debtor and case unknown")
        if mismatch: hold.append(f"printed name ≠ registry name for the printed CNPJ: {mismatch} — confirm who the creditor is")
        rec = 0.0 if hold else (plan[0] if plan else BAND_MAP[band][0] * (DEFAULT_FIN_MULT if fin else 1.0))
        fund = float(face_d) * rec / (1 + DEFAULT_IRR) ** _years(band, ysf, plan)
        # --- contacts: Apollo first (ranked), RFB partners (QSA) as fallback; a non-finance Apollo pick yields the DM slot to a registered partner ---
        ap = db.execute("SELECT person_name, title, email, mobile, direct_phone, email_status, pick_note FROM apollo_contacts WHERE cnpj_basico=? ORDER BY rank", (root,)).fetchall()
        rf = db.execute("SELECT person_name, role FROM contacts WHERE cnpj_basico=? ORDER BY CASE role_code WHEN 49 THEN 0 WHEN 16 THEN 1 WHEN 5 THEN 2 WHEN 10 THEN 3 ELSE 9 END, person_name", (root,)).fetchall()
        empty = {"name": "", "role": "", "mobile": "", "phone": "", "email": "", "status": "", "note": "", "src": "", "tier": ""}
        people = [{"name": a[0] or "", "role": a[1] or "", "mobile": a[3] or "", "phone": a[4] or "", "email": a[2] or "", "status": a[5] or "", "note": a[6] or "", "src": "apollo", "tier": _tier(a[1])} for a in ap if a[0]]
        rfb_people = [{**empty, "name": n, "role": f"{ro} (RFB partner)", "src": "rfb", "tier": "A"} for n, ro in rf]
        if people and people[0]["tier"] == "C" and rfb_people:
            people = [rfb_people[0]] + people; rfb_people = rfb_people[1:]
        ordered = people + rfb_people
        primary = ordered[0] if ordered else dict(empty)
        backup = next((p for p in ordered[1:] if p["email"] or p["src"] == "rfb" or p["phone"]), ordered[1] if len(ordered) > 1 else dict(empty))
        notes = []
        if primary["src"] == "apollo":
            if primary["tier"] == "C": notes.append(f"DM is not a finance role ({primary['role'][:30]})")
            if primary["note"]: notes.append("DM: " + primary["note"])
            if not primary["email"]: notes.append("no work email in Apollo for DM")
        elif primary["src"] == "rfb":
            notes.append("DM = registered partner (RFB QSA)" + ("; Apollo pick demoted to backup (non-finance role)" if people and people[0]["src"] == "rfb" and len(people) > 1 else "; no Apollo match"))
        else:
            notes.append("no named contact found")
        if backup["src"] == "apollo" and backup["note"]: notes.append("BACKUP: " + backup["note"])
        if backup["src"] == "apollo" and backup["email"] and backup["status"] == "extrapolated": notes.append("backup email is pattern-guessed")
        if backup["phone"] and _phone(backup["phone"]) in (_phone(phone), _phone(phone2)): backup = {**backup, "phone": ""}
        no_route = not any([_phone(phone), _phone(phone2), primary["email"], primary["mobile"], backup["email"], backup["phone"]])
        sstatus = _seller_status(razao, name_printed, situacao, situacao_desc)
        if sstatus not in ("ATIVA",): notes.append("SELLER: " + sstatus)
        if not razao: flags = (flags + " NO_RFB_RECORD").strip()
        if no_route: flags = (flags + " NO_ROUTE_IN").strip(); notes.append("NO ROUTE IN: no phone or email on file")
        extra_flags = []
        if two_values: extra_flags.append("TWO VALUES PRINTED — confirm which figure the creditor recognises")
        if share is not None and share >= 0.5: extra_flags.append(f"CONTROL POSITION: {share:.0%} of class III by value — carries the class vote (art. 45 §1)")
        elif share is not None and share > 1 / 3: extra_flags.append(f"BLOCKING POSITION: {share:.0%} of class III — can defeat a cram-down (art. 58 §1)")
        if mismatch: extra_flags.append("VERIFY creditor identity")
        if extra_share_note: extra_flags.append(extra_share_note)
        if cred_in_rj and not rel: extra_flags.append("seller is in RJ (art. 66)")
        court_txt = court if court and court != "UNKNOWN" else (("TJ" + TR.get(case[18:20], "")) if case else "")
        stage_txt = {"ACTIVE": "DataJud: live (no closure/falência movement)", "CONVERTED_TO_BANKRUPTCY": "DataJud: falência decreed", "CLOSED": "DataJud: closed/archived"}.get(
            stage or dstage or "", stage or dstage or ("not in DataJud (RJ-class pull only)" if case else ""))
        plan_ev = {"CONFIRMED": "GRANT DOC FOUND (art. 58)", "PLAN_FILED": "PLAN FILED (art. 53), not granted", "ESTIMATED": "NONE FOUND", "UNKNOWN": "NONE FOUND"}.get(pstatus or "UNKNOWN", pstatus or "")
        if granted: plan_ev = "GRANT in DataJud (art. 58)"
        lead = {"root": root, "case": case, "face": face_d, "n": ncl, "name_printed": name_printed, "razao": razao, "phone": phone, "phone2": phone2,
                "uf": uf, "city": city, "cnae": cnae, "porte": porte, "capital": capital, "url": url, "dtype": kind,
                "pub": pub, "proof": proof, "court": court_txt, "stage": stage_txt, "year": year, "ysf": ysf, "debtor": debtor,
                "debtor_cnpj": vcnpj or accepted_cnpj or "", "band": band, "breason": breason or "", "pstatus": plan_ev,
                "fund": fund, "primary": primary, "backup": backup, "note": "; ".join(notes), "claims": claims, "vals_printed": vals_printed,
                "flags": " · ".join([FLAG_TEXT.get(f, f) for f in flags.replace("|", " ").split() if not f.startswith(("ROWS:", "VALUE_COL:"))] + extra_flags),
                "fit": _seller_fit(razao, capital, porte), "book": book, "ctype": ctype, "sstatus": sstatus, "plan": plan, "hold": "; ".join(hold),
                "ests": ests_l, "headings": headings, "recup_rows": recup_rows, "share": share, "iii_total": iii_total,
                "dsource": DEBTOR_SOURCE.get((dsource if ddisplay else None) or (None if not debtor else ("rfb" if debtor == drazao else "hint")), "not identified"),
                "proceeding": PROCEEDING.get(proc or "UNKNOWN", proc) or ("Recuperação judicial (from the list type; head not read)" if case else "")}
        out.append(lead)
    out.sort(key=lambda x: (-x["fund"], -float(x["face"])))
    return out, excluded


def _assumptions(ws, hidden=False):
    ws.title = "Assumptions"
    ws.column_dimensions["A"].width = 14; ws.column_dimensions["F"].width = 100
    ws["A1"] = "PRICING ASSUMPTIONS — yellow cells are editable; every quote on every call tab recalculates. THIS IS A MODEL, NOT A PRICE."; ws["A1"].font = Font(bold=True, size=12)
    ws["A3"] = "Target investor IRR (annual)"; ws["B3"] = DEFAULT_IRR; ws["B3"].number_format = "0%"
    ws["A4"] = "Our buy price as share of fund price"; ws["B4"] = DEFAULT_SHARE; ws["B4"].number_format = "0%"
    ws["A5"] = "Financial-creditor recovery multiplier"; ws["B5"] = DEFAULT_FIN_MULT; ws["B5"].number_format = "0.00"
    ws["C5"] = "× band recovery for every creditor whose type says FINANCIAL (banks, FIDCs, co-ops on any tab, state banks included); not applied when a plan's own terms were read. 1.00 = same as suppliers"
    for c in ("B3", "B4", "B5"): ws[c].fill = PatternFill("solid", fgColor="FFF2CC")
    for c, h in zip("ABCDEF", ("Debtor band", "Class III recovery %", "Grace (y)", "Term (y)", "Plan lag (y)", "What the band means")):
        ws[f"{c}7"] = h; ws[f"{c}7"].font = WHITE; ws[f"{c}7"].fill = HEAD
    for i, (b, rec, g, t, lag, note) in enumerate(BANDS, start=8):
        ws[f"A{i}"], ws[f"B{i}"], ws[f"C{i}"], ws[f"D{i}"], ws[f"E{i}"], ws[f"F{i}"] = b, rec, g, t, lag, note
        ws[f"B{i}"].number_format = "0%"
        for c in "BCDE": ws[f"{c}{i}"].fill = PatternFill("solid", fgColor="FFF2CC")
    lines = [
        "Fund price = face × recovery% ÷ (1+IRR)^(years to payment). Our quote = fund price × share. Our ágio = fund price − quote.",
        "A row with text in PRICING HOLD prices at zero on purpose: the source is not a statutory list, the case is not identified, or the printed name does not match the CNPJ. Clear the hold (delete the text) once the point is verified and the row prices like any other.",
        "Where a plan PDF was read (columns 'Plan: class III recovery / grace / term'), the plan's own class III general condition replaces the band's recovery, grace and term, and the financial multiplier is not applied on top.",
        "Years to payment — band A (recovery granted, art. 58): max(1, grace + term/2 − years already elapsed under the plan), where the grant is assumed 'plan lag' years after filing.",
        "Years to payment — bands B, C, U: plan lag (years from today until a plan is granted) + grace + term/2. Band D is not priced: those rows sit on EXCLUDED (review).",
        "So an old case with no granted plan prices LOWER than a fresh one, never higher. Band recovery % is an assumption, not a term read from any plan, unless the Plan columns are filled.",
        "Terminology: a plan is FILED by the debtor (art. 53), APPROVED by the creditors (AGC, art. 45) and the recovery is GRANTED by the court (art. 58). Only the grant starts the payment clock.",
        "Financial creditors (banks, FIDCs, co-ops) usually get worse class III terms than trade suppliers; the multiplier in B5 haircuts their recovery wherever the 'Creditor type' column says FINANCIAL.",
        "A cessão de crédito is not done when the creditor says yes: it needs a written instrument (art. 288 CC, registered to bind third parties), notification of the debtor (art. 290 CC), and the assignee's substitution in the creditor list by the AJ (art. 7/18 LRF). The STATUS chain on the call tabs follows those steps. A seller that is itself in RJ needs its own judge's authorisation (art. 66).",
        "Everything on this tab is a MODEL. The face value, the page/row and the source document are court data; the quote is ours.",
    ]
    for i, t in enumerate(lines, start=14): ws[f"A{i}"] = t
    if hidden: ws.sheet_state = "hidden"


def _link(cell, url):
    if not url: return
    cell.hyperlink = quote(url, safe=":/?&=%#+~;@!$,()'*"); cell.font = Font(color="0563C1", underline="single")


def _date(s):
    try: return datetime.strptime(str(s)[:10], "%Y-%m-%d").date() if s else ""
    except ValueError: return s or ""


def _call_sheet(ws, leads, title="CALL SHEET"):
    ws.title = title
    ws.freeze_panes = "D2"
    for j, (h, w, g) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=j, value=h); c.font = WHITE; c.fill = HEAD; c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 34
    dv = DataValidation(type="list", formula1='"' + ",".join(STATUSES) + '"', allow_blank=True, showErrorMessage=True, errorStyle="stop", error="Pick a status from the list"); ws.add_data_validation(dv)
    dvb = DataValidation(type="list", formula1='"A,B,C,D,U"', allow_blank=True, showErrorMessage=True, errorStyle="stop", error="Band must be A, B, C, D or U"); ws.add_data_validation(dvb)
    thin = Side(style="thin", color="D9D9D9")
    VL = lambda band, k, dflt: f"IFERROR(VLOOKUP({band},Assumptions!$A$8:$F$12,{k},FALSE),{dflt})"
    for i, Ld in enumerate(leads, start=2):
        r = i; face = f"{L('FACE VALUE (as printed)')}{r}"; band = f"{L('DEBTOR BAND')}{r}"; filed = f"{L('Filed')}{r}"; hold = f"{L('PRICING HOLD (why unpriced)')}{r}"
        prec = f"{L('Plan: class III recovery')}{r}"; pgr = f"{L('Plan: grace (y)')}{r}"; pterm = f"{L('Plan: term (y)')}{r}"; ctype = f"{L('Creditor type')}{r}"
        rec = f"{L('Exp. recovery %')}{r}"; yrs = f"{L('Years to payment')}{r}"; fund = f"{L('FUND PRICE @ IRR (model, R$)')}{r}"; quote_ = f"{L('OUR QUOTE (model, R$)')}{r}"
        elapsed = f'IF({filed}="",0,YEAR(TODAY())-{filed})'
        grace = f'IF({pgr}<>"",{pgr},{VL(band, 3, 3)})'; term = f'IF({pterm}<>"",{pterm},{VL(band, 4, 8)})'; lag = VL(band, 5, 2.5)
        mult = f'IF(ISNUMBER(SEARCH("FINANCIAL",{ctype})),Assumptions!$B$5,1)'
        p, b = Ld["primary"], Ld["backup"]; plan = Ld["plan"]
        vals = {
            "CREDITOR — company owed money": Ld["razao"] or Ld["name_printed"], "AMOUNT OWED (R$)": float(Ld["face"]), "DEBTOR — party in RJ": Ld["debtor"] or "(debtor not identified — see proof)",
            "STATUS": "NEW", "LAST TOUCH": "", "NEXT ACTION": "", "NEXT DATE": "", "NOTES": "",
            "DECISION MAKER": p["name"], "Role": p["role"], "Company phone 1": _phone(Ld["phone"]), "Company phone 2": _phone(Ld["phone2"]), "Mobile (Apollo; reveal pending)": _phone(p["mobile"]),
            "Email": p["email"], "Email status": p["status"] if p["email"] else "",
            "BACKUP CONTACT": b["name"], "Backup role": b["role"], "Backup phone": _phone(b["phone"]) if b["phone"] else "", "Backup email": b["email"], "Contact note": Ld["note"],
            "FACE VALUE (as printed)": float(Ld["face"]), "DEBTOR BAND": Ld["band"], "PRICING HOLD (why unpriced)": Ld["hold"],
            "Plan: class III recovery": plan[0] if plan else "", "Plan: grace (y)": plan[1] if plan else "", "Plan: term (y)": plan[2] if plan else "",
            "Exp. recovery %": f'=IF({hold}<>"",0,IF({prec}<>"",{prec},{VL(band, 2, 0)}*{mult}))',
            "Years to payment": f'=IF({band}="A",MAX(1,{grace}+{term}/2-MAX(0,{elapsed}-{VL(band, 5, 1.5)})),{lag}+{grace}+{term}/2)',
            "FUND PRICE @ IRR (model, R$)": f"={face}*{rec}/(1+Assumptions!$B$3)^{yrs}",
            "Fund ¢/R$": f'=IF({face}=0,"",ROUND({fund}/{face}*100,1))',
            "OUR QUOTE (model, R$)": f"={fund}*Assumptions!$B$4",
            "Quote ¢/R$": f'=IF({face}=0,"",ROUND({quote_}/{face}*100,1))',
            "OUR ÁGIO (model, R$)": f"={fund}-{quote_}", "Ágio pts of face": f'=IF({face}=0,"",ROUND(({fund}-{quote_})/{face}*100,1))',
            "PROOF QUALITY": Ld["proof"], "PROOF: source document": Ld["url"] or "", "Page / row (extractor index)": "; ".join(f"p{pg} r{ri}" for pg, ri, *_ in Ld["claims"]),
            "Value as printed": Ld["vals_printed"], "Creditor name as printed": Ld["name_printed"] or "", "Creditor CNPJ(s) as printed": "; ".join(_cnpj(e) for e in Ld["ests"] if e),
            "Printed section heading": " | ".join(dict.fromkeys(h[:70] for h in Ld["headings"])), "Recuperanda on the row (if printed)": "; ".join(Ld["recup_rows"]),
            "Share of class III on this list": (Ld["share"] if Ld["share"] is not None else ""), "Class III total on list (R$)": (float(Ld["iii_total"]) if Ld["iii_total"] is not None else ""),
            "Which list": Ld["dtype"], "List date": _date(Ld["pub"]), "Plan evidence": Ld["pstatus"],
            "Case number": Ld["case"] or "(not printed)", "Court": Ld["court"], "Proceeding": Ld["proceeding"], "Stage (DataJud)": Ld["stage"], "Filed": Ld["year"] or "", "Band reason": Ld["breason"],
            "Debtor CNPJ (lead entity)": _cnpj(Ld["debtor_cnpj"]), "Debtor source": Ld["dsource"], "Plan terms source": (plan[3] if plan else ""),
            "Creditor type": Ld["ctype"], "Seller status": Ld["sstatus"],
            "City": Ld["city"] or "", "UF": Ld["uf"] or "", "Sector": Ld["cnae"] or "", "Size": {"DEMAIS": "não ME/EPP"}.get(Ld["porte"] or "", Ld["porte"] or ""),
            "Capital (R$)": (float(Ld["capital"]) if Ld["capital"] not in (None, "") else ""), "Seller fit": Ld["fit"],
            "CNPJ root": Ld["root"], "Claims": Ld["n"], "Flags": Ld["flags"],
        }
        for j, (h, w, g) in enumerate(COLS, 1):
            c = ws.cell(row=r, column=j, value=vals.get(h))
            if g in FILL: c.fill = PatternFill("solid", fgColor=FILL[g])
            c.border = Border(bottom=thin)
        for h in ("AMOUNT OWED (R$)", "FACE VALUE (as printed)", "FUND PRICE @ IRR (model, R$)", "OUR QUOTE (model, R$)", "OUR ÁGIO (model, R$)", "Capital (R$)", "Class III total on list (R$)"):
            ws.cell(row=r, column=C[h]).number_format = "#,##0"
        for h in ("Exp. recovery %", "Plan: class III recovery", "Share of class III on this list"): ws.cell(row=r, column=C[h]).number_format = "0%"
        ws.cell(row=r, column=C["Years to payment"]).number_format = "0.0"
        ws.cell(row=r, column=C["OUR QUOTE (model, R$)"]).font = Font(bold=True); ws.cell(row=r, column=C["OUR ÁGIO (model, R$)"]).font = Font(bold=True)
        ws.cell(row=r, column=C["CREDITOR — company owed money"]).font = Font(bold=True)
        bc = ws.cell(row=r, column=C["DEBTOR BAND"]); bc.fill = PatternFill("solid", fgColor=BAND_FILL.get(Ld["band"], "EDEDED")); bc.alignment = Alignment(horizontal="center"); bc.font = Font(bold=True)
        if Ld["hold"]: ws.cell(row=r, column=C["PRICING HOLD (why unpriced)"]).fill = PatternFill("solid", fgColor="FF9999")
        pq = ws.cell(row=r, column=C["PROOF QUALITY"])
        for k, colr in PROOF_FILL.items():
            if Ld["proof"].startswith(k): pq.fill = PatternFill("solid", fgColor=colr)
        if not Ld["debtor"]: ws.cell(row=r, column=C["DEBTOR — party in RJ"]).font = Font(italic=True, color="9C0006")
        if p["src"] == "rfb": ws.cell(row=r, column=C["Role"]).fill = PatternFill("solid", fgColor="EDEDED")
        _link(ws.cell(row=r, column=C["PROOF: source document"]), Ld["url"])
        for h in ("LAST TOUCH", "NEXT DATE", "List date"): ws.cell(row=r, column=C[h]).number_format = "yyyy-mm-dd"
    n = len(leads) + 1
    dv.add(f"{L('STATUS')}2:{L('STATUS')}{max(2, n)}"); dvb.add(f"{L('DEBTOR BAND')}2:{L('DEBTOR BAND')}{max(2, n)}")
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{max(2, n)}"
    S = L("STATUS")
    for s, colr in (("SETTLED", "C6EFCE"), ("CESSÃO SIGNED", "C6EFCE"), ("SUBSTITUTED IN LIST", "C6EFCE"), ("TERMS AGREED", "E2EFDA"), ("INTERESTED", "E2EFDA"), ("DEAD", "FFC7CE"), ("NOT INTERESTED", "FFC7CE")):
        ws.conditional_formatting.add(f"{S}2:{S}{n}", CellIsRule(operator="equal", formula=[f'"{s}"'], fill=PatternFill("solid", fgColor=colr)))
    ND = L("NEXT DATE")
    ws.conditional_formatting.add(f"{ND}2:{ND}{n}", FormulaRule(formula=[f'AND({ND}2<>"",{ND}2<TODAY(),{S}2<>"DEAD",{S}2<>"SETTLED")'], fill=PatternFill("solid", fgColor="FFC7CE")))
    ws.conditional_formatting.add(f"{ND}2:{ND}{n}", FormulaRule(formula=[f'{ND}2=TODAY()'], fill=PatternFill("solid", fgColor="FFEB9C")))
    return n


def _excluded(ws, rows):
    ws.title = "EXCLUDED (review)"
    heads = ["CREDITOR", "DEBTOR", "FACE (R$)", "WHY EXCLUDED", "Would-be tab", "Case number", "CNPJ root", "Source document"]
    for j, h in enumerate(heads, 1):
        c = ws.cell(row=1, column=j, value=h); c.font = WHITE; c.fill = HEAD
    for w, col in zip((42, 40, 16, 90, 14, 26, 10, 60), "ABCDEFGH"): ws.column_dimensions[col].width = w
    for i, x in enumerate(sorted(rows, key=lambda x: -float(x["face"])), start=2):
        ws.cell(row=i, column=1, value=x["razao"] or x["name_printed"]); ws.cell(row=i, column=2, value=x["debtor"]); ws.cell(row=i, column=3, value=float(x["face"])).number_format = "#,##0"
        ws.cell(row=i, column=4, value=x["reason"]); ws.cell(row=i, column=5, value=x["book"]); ws.cell(row=i, column=6, value=x["case"] or "(not printed)"); ws.cell(row=i, column=7, value=x["root"])
        c = ws.cell(row=i, column=8, value=x["url"] or ""); _link(c, x["url"])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{max(2, len(rows) + 1)}"


STACK = [("CREDITOR — company owed money", "text"), ("DEBTOR — party in RJ", "text"), ("FACE VALUE (as printed)", "num"),
         ("FUND PRICE @ IRR (model, R$)", "num"), ("OUR ÁGIO (model, R$)", "num"), ("DEBTOR BAND", "text"), ("STATUS", "text"),
         ("NEXT DATE", "date"), ("UF", "text"), ("Seller fit", "text"), ("Email", "text"), ("PROOF QUALITY", "text"), ("Case number", "text"), ("Flags", "text"),
         ("PRICING HOLD (why unpriced)", "text")]
SC = {h: get_column_letter(i + 1) for i, (h, _) in enumerate(STACK)}
SC["BOOK"] = get_column_letter(len(STACK) + 1)
SC["Where"] = get_column_letter(len(STACK) + 2)
SC["KEY"] = get_column_letter(len(STACK) + 3)   # ágio + a row-unique epsilon (above double-precision ULP of any ágio, far below R$0.01), so LARGE/MATCH never return the same row twice


def _book(ws, books):
    """Hidden sheet stacking every call tab by direct cell reference, so the dashboard measures the whole book, not one tab."""
    ws.title = "_BOOK"
    for j, h in enumerate([h for h, _ in STACK] + ["BOOK", "Where", "ÁGIO KEY"], 1):
        ws.cell(row=1, column=j, value=h).font = Font(bold=True)
    r = 1
    for title, book, last in books:
        for i in range(2, last + 1):
            r += 1
            for j, (h, kind) in enumerate(STACK, 1):
                ref = f"'{title}'!${L(h)}${i}"
                ws.cell(row=r, column=j, value=f"={ref}" if kind == "num" else f'=IF({ref}="","",{ref})')
            ws.cell(row=r, column=len(STACK) + 1, value=book)
            ws.cell(row=r, column=len(STACK) + 2, value=f"{title} · row {i}")
            ws.cell(row=r, column=len(STACK) + 3, value=f"={SC['OUR ÁGIO (model, R$)']}{r}+ROW()*0.000001")
    ws.sheet_state = "hidden"
    return r


def _dashboard(ws, leads, nrows, books, stamp):
    ws.title = "DASHBOARD"
    rng = lambda h: f"_BOOK!${SC[h]}$2:${SC[h]}${nrows}"
    A = rng("OUR ÁGIO (model, R$)"); F = rng("FACE VALUE (as printed)"); S = rng("STATUS"); BK = rng("BOOK"); K = rng("KEY"); HO = rng("PRICING HOLD (why unpriced)")
    ws.column_dimensions["A"].width = 44; ws.column_dimensions["B"].width = 16; ws.column_dimensions["C"].width = 16; ws.column_dimensions["D"].width = 4
    for col, w in zip("EFGHIJK", (36, 30, 15, 7, 16, 12, 28)): ws.column_dimensions[col].width = w
    ws["A1"] = "ANTEPARO — CALL DASHBOARD"; ws["A1"].font = Font(bold=True, size=16, color=NAVY)
    ws["A2"] = f"All call tabs counted together · {stamp}"; ws["A2"].font = Font(italic=True, color="666666")
    ws["A3"] = "Fund price, quote and ágio are a MODEL off the Assumptions tab (IRR, band recovery, plan clock). Face value, page/row and the proof link are court data."
    ws["A3"].font = Font(italic=True, color="9C0006")

    def block(row, title, cols):
        ws[f"A{row}"] = title
        for c, lab in zip("BC", cols): ws[f"{c}{row}"] = lab
        for c in "ABC": ws[f"{c}{row}"].font = WHITE; ws[f"{c}{row}"].fill = HEAD
        return row + 1

    pipe = "+".join(f'SUMIF({S},"{s}",{A})' for s in PIPELINE)
    kpis = [("Leads in book (all tabs)", f'=SUMPRODUCT(--({rng("CREDITOR — company owed money")}<>""))', "0"),
            ("Total face (R$)", f"=SUM({F})", "#,##0"),
            ("Model fund price pipeline (R$)", f'=SUM({rng("FUND PRICE @ IRR (model, R$)")})', "#,##0"),
            ("Model ÁGIO pipeline (R$)", f"=SUM({A})", "#,##0"),
            ("Rows on PRICING HOLD (zero quote until cleared)", f'=SUMPRODUCT(--({HO}<>""))', "0"),
            ("Face on pricing hold (R$)", f'=SUMIF({HO},"<>",{F})', "#,##0"),
            ("Untouched (NEW)", f'=COUNTIF({S},"NEW")', "0"),
            ("Calls due today", f'=COUNTIF({rng("NEXT DATE")},TODAY())', "0"),
            ("Overdue follow-ups", f'=COUNTIFS({rng("NEXT DATE")},"<"&TODAY(),{rng("NEXT DATE")},"<>",{S},"<>DEAD",{S},"<>SETTLED")', "0"),
            ("Ágio in progress (INTERESTED → SUBSTITUTED)", f"={pipe}", "#,##0"),
            ("Ágio realised (SETTLED)", f'=SUMIF({S},"SETTLED",{A})', "#,##0"),
            ("Rows with a decision-maker email", f'=SUMPRODUCT(--({rng("Email")}<>""))', "0"),
            ("Rows with no route in (no phone/email)", f'=COUNTIF({rng("Flags")},"*no route*")', "0"),
            ("Rows needing a proof check", f'=COUNTIF({rng("PROOF QUALITY")},"NO TOTALS*")+COUNTIF({rng("PROOF QUALITY")},"PARTIAL*")+COUNTIF({rng("PROOF QUALITY")},"CHECK*")', "0"),
            ("Ágio on rows needing a proof check (R$)", f'=SUMIF({rng("PROOF QUALITY")},"NO TOTALS*",{A})+SUMIF({rng("PROOF QUALITY")},"PARTIAL*",{A})+SUMIF({rng("PROOF QUALITY")},"CHECK*",{A})', "#,##0")]
    r = block(5, "PIPELINE", ("", ""))
    for lab, f, fmt in kpis:
        ws[f"A{r}"] = lab; ws[f"B{r}"] = f; ws[f"B{r}"].number_format = fmt; ws[f"B{r}"].font = Font(bold=True, size=12); r += 1

    r = block(r + 1, "BOOK", ("leads", "ágio (R$)"))
    for lab, key in (("TRADE — CALL SHEET tab", "TRADE"), ("FINANCIAL — banks, FIDCs, co-ops", "FINANCIAL"), ("STATE-OWNED — public banks & companies", "STATE-OWNED")):
        ws[f"A{r}"] = lab; ws[f"B{r}"] = f'=COUNTIF({BK},"{key}")'; ws[f"C{r}"] = f'=SUMIF({BK},"{key}",{A})'; ws[f"C{r}"].number_format = "#,##0"; r += 1
    ws[f"A{r}"] = "All books"; ws[f"B{r}"] = f'=SUMPRODUCT(--({BK}<>""))'; ws[f"C{r}"] = f"=SUM({A})"; ws[f"C{r}"].number_format = "#,##0"
    for c in "ABC": ws[f"{c}{r}"].font = Font(bold=True)
    r += 1

    r = block(r + 1, "FUNNEL", ("leads", "ágio (R$)"))
    for s in STATUSES:
        ws[f"A{r}"] = s; ws[f"B{r}"] = f'=COUNTIF({S},"{s}")'; ws[f"C{r}"] = f'=SUMIF({S},"{s}",{A})'; ws[f"C{r}"].number_format = "#,##0"; r += 1

    r = block(r + 1, "DEBTOR BAND", ("leads", "ágio (R$)"))
    B = rng("DEBTOR BAND")
    for b, *_ in BANDS:
        ws[f"A{r}"] = BAND_SHORT[b]; ws[f"B{r}"] = f'=COUNTIF({B},"{b}")'; ws[f"C{r}"] = f'=SUMIF({B},"{b}",{A})'
        ws[f"C{r}"].number_format = "#,##0"; ws[f"A{r}"].fill = PatternFill("solid", fgColor=BAND_FILL[b]); r += 1

    r = block(r + 1, "SELLER FIT", ("leads", "ágio (R$)"))
    F2 = rng("Seller fit")
    for fit in ("MID-MARKET", "SMALL", "LARGE CORP", "UNKNOWN"):
        ws[f"A{r}"] = fit + (" — sells through a desk, if at all" if fit == "LARGE CORP" else (" — no RFB record" if fit == "UNKNOWN" else "")); ws[f"B{r}"] = f'=COUNTIF({F2},"{fit}")'
        ws[f"C{r}"] = f'=SUMIF({F2},"{fit}",{A})'; ws[f"C{r}"].number_format = "#,##0"; r += 1

    r = block(r + 1, "PROOF QUALITY", ("leads", "ágio (R$)"))
    PQ = rng("PROOF QUALITY")
    for lab, pat in (("OK — reconciled against printed totals", "OK*"), ("NO TOTALS — list prints none", "NO TOTALS*"), ("PARTIAL — class III reconciled only", "PARTIAL*"), ("CHECK — not a statutory list / class III did not reconcile", "CHECK*")):
        ws[f"A{r}"] = lab; ws[f"B{r}"] = f'=COUNTIF({PQ},"{pat}")'; ws[f"C{r}"] = f'=SUMIF({PQ},"{pat}",{A})'; ws[f"C{r}"].number_format = "#,##0"; r += 1

    notes = ["How to use: work each call tab top-down (sorted by model ágio). Set STATUS and NEXT DATE on every touch; this tab updates itself. Do not delete rows (mark DEAD); sorting and filtering are fine.",
             "STATUS chain after a yes: TERMS AGREED → CESSÃO SIGNED (written instrument, art. 288 CC) → SUBSTITUTED IN LIST (debtor notified, art. 290 CC; AJ updates the list) → SETTLED (fund pays). Ágio counts as in-progress until then.",
             "Bands: A recovery granted (art. 58) · B live case filed ≤2 y ago, plan pending · C older case, grant not seen · U unresolved (on hold). Band D (falência / extinguished / closed) is not in the book: see EXCLUDED (review).",
             "PRICING HOLD: rows kept in the book at a zero quote because the source is not a statutory list, the case is not identified, or the printed name does not match the printed CNPJ. Clear the hold text once verified.",
             "PROOF QUALITY: OK = the list's own printed totals reconcile; NO TOTALS = the list prints none, so the value stands on its row alone; PARTIAL = class III reconciled but another class did not; CHECK = not an art. 7/51 list, or the class III total did not reconcile.",
             "FINANCIAL CREDITORS = banks, FIDCs, securitizers, credit co-ops: same class III paper, sold through a desk; plans usually give them worse terms (Assumptions B5, applied wherever Creditor type says FINANCIAL). STATE-OWNED = public banks and state companies: sell by auction/desk only, if at all.",
             "Flags to read before dialling: CONTROL / BLOCKING POSITION (the creditor's share of class III on that list), TWO VALUES PRINTED, VERIFY creditor identity, seller is in RJ (art. 66). 'Seller status' says who can actually sign for the seller.",
             "EXCLUDED (review) lists rows kept out of the book and why (intercompany / group member, subordinated or class I-II-IV headings, ME/EPP creditors, falência or extinguished proceedings, prose lines naming another CNPJ, public bodies, inactive creditors).",
             "Mobile numbers are empty: Apollo direct-dial credits reset on 21 Sep 2026; company switchboards and work emails are filled where found. 'Debtor source' says whether the debtor name was printed on the list (read twice) or taken from the portal page."]
    for i, t in enumerate(notes, start=r + 1): ws[f"A{i}"] = t

    ws["E5"] = "TOP 15 BY MODEL ÁGIO"; ws["F5"] = "Debtor"; ws["G5"] = "Ágio (R$)"; ws["H5"] = "Band"; ws["I5"] = "Status"; ws["J5"] = "Next date"; ws["K5"] = "Where to find it"
    for c in "EFGHIJK": ws[f"{c}5"].font = WHITE; ws[f"{c}5"].fill = HEAD
    for k in range(1, 16):
        rr = 5 + k
        key = f"IFERROR(LARGE({K},{k}),\"\")"
        mrow = f"MATCH({key},{K},0)"
        ws[f"G{rr}"] = f'=IFERROR(INDEX({A},{mrow}),"")'; ws[f"G{rr}"].number_format = "#,##0"
        for col, h in (("E", "CREDITOR — company owed money"), ("F", "DEBTOR — party in RJ"), ("H", "DEBTOR BAND"), ("I", "STATUS"), ("J", "NEXT DATE"), ("K", "Where")):
            ws[f"{col}{rr}"] = f'=IFERROR(INDEX({rng(h)},{mrow}),"")'
        ws[f"J{rr}"].number_format = "yyyy-mm-dd"

    cnt = Counter(Ld["case"] for Ld in leads if Ld["case"]); agio = defaultdict(float); name = {}
    for Ld in leads:
        if Ld["case"]: agio[Ld["case"]] += Ld["fund"] * (1 - DEFAULT_SHARE); name.setdefault(Ld["case"], Ld["debtor"] or "(debtor not identified)")
    top = sorted(cnt.items(), key=lambda kv: (-kv[1], -agio[kv[0]]))[:12]
    r2 = 22
    ws[f"E{r2}"] = "CASE OWNERSHIP — deepest cases"; ws[f"F{r2}"] = "case"; ws[f"G{r2}"] = "leads"; ws[f"H{r2}"] = ""; ws[f"I{r2}"] = "ágio (R$)"; ws[f"J{r2}"] = "called"; ws[f"K{r2}"] = "in progress"
    for c in "EFGHIJK": ws[f"{c}{r2}"].font = WHITE; ws[f"{c}{r2}"].fill = HEAD
    CN = rng("Case number")
    for i, (case, _) in enumerate(top, start=r2 + 1):
        ws[f"E{i}"] = name[case]; ws[f"F{i}"] = case; ws[f"G{i}"] = f'=COUNTIF({CN},"{case}")'; ws[f"I{i}"] = f'=SUMIF({CN},"{case}",{A})'; ws[f"I{i}"].number_format = "#,##0"
        ws[f"J{i}"] = f'=COUNTIFS({CN},"{case}",{S},"<>NEW")'; ws[f"K{i}"] = "=" + "+".join(f'COUNTIFS({CN},"{case}",{S},"{s}")' for s in PIPELINE)
    r3 = r2 + len(top) + 2
    ws[f"E{r3}"] = "BY CREDITOR STATE — top 10 by leads (RFB seat)"; ws[f"F{r3}"] = "leads"; ws[f"G{r3}"] = "ágio (R$)"
    for c in "EFG": ws[f"{c}{r3}"].font = WHITE; ws[f"{c}{r3}"].fill = HEAD
    U = rng("UF")
    top_uf = [uf for uf, _ in Counter(Ld["uf"] for Ld in leads if Ld["uf"]).most_common(10)]
    i = r3
    for i, uf in enumerate(top_uf, start=r3 + 1):
        ws[f"E{i}"] = uf; ws[f"F{i}"] = f'=COUNTIF({U},"{uf}")'; ws[f"G{i}"] = f'=SUMIF({U},"{uf}",{A})'; ws[f"G{i}"].number_format = "#,##0"
    others = "-".join(f'COUNTIF({U},"{uf}")' for uf in top_uf); others_a = "-".join(f'SUMIF({U},"{uf}",{A})' for uf in top_uf)
    ws[f"E{i + 1}"] = "other states"; ws[f"F{i + 1}"] = f'=SUMPRODUCT(--({U}<>""))-{others}'; ws[f"G{i + 1}"] = f'=SUMIF({U},"<>",{A})-{others_a}'; ws[f"G{i + 1}"].number_format = "#,##0"
    ws[f"E{i + 2}"] = "no RFB seat (no RFB record)"; ws[f"F{i + 2}"] = f'=SUMPRODUCT(--({U}=""),--({rng("CREDITOR — company owed money")}<>""))'; ws[f"G{i + 2}"] = f'=SUMPRODUCT(--({U}=""),{A})'; ws[f"G{i + 2}"].number_format = "#,##0"
    ws[f"E{i + 3}"] = "all rows"; ws[f"F{i + 3}"] = f'=SUM(F{r3 + 1}:F{i + 2})'; ws[f"G{i + 3}"] = f'=SUM(G{r3 + 1}:G{i + 2})'; ws[f"G{i + 3}"].number_format = "#,##0"
    for c in "EFG": ws[f"{c}{i + 3}"].font = Font(bold=True)


def _stamp(run_id):
    try: sha = subprocess.run(["git", "rev-parse", "--short", "HEAD"], capture_output=True, text=True, timeout=5).stdout.strip()
    except Exception: sha = ""
    return f"run {run_id} · build {sha or 'uncommitted'} · {datetime.utcnow().strftime('%Y-%m-%d %H:%MZ')}"


def export_callsheet(db, run_id, path, hide_assumptions=False):
    every, excluded = _leads(db, run_id)
    by = {"TRADE": [l for l in every if l["book"] == "TRADE"], "FINANCIAL": [l for l in every if l["book"] == "FINANCIAL"], "STATE-OWNED": [l for l in every if l["book"] == "STATE-OWNED"]}
    wb = Workbook()
    dash = wb.active
    books = [("CALL SHEET", "TRADE", _call_sheet(wb.create_sheet(), by["TRADE"]))]
    if by["FINANCIAL"]:
        books.append(("FINANCIAL CREDITORS", "FINANCIAL", _call_sheet(wb.create_sheet(), by["FINANCIAL"], title="FINANCIAL CREDITORS")))
    if by["STATE-OWNED"]:
        books.append(("STATE-OWNED CREDITORS", "STATE-OWNED", _call_sheet(wb.create_sheet(), by["STATE-OWNED"], title="STATE-OWNED CREDITORS")))
    _excluded(wb.create_sheet(), excluded)
    nrows = _book(wb.create_sheet(), books)
    _dashboard(dash, every, nrows, books, _stamp(run_id))
    _assumptions(wb.create_sheet(), hidden=hide_assumptions)
    wb.save(path)
    tally = lambda rows: {"n": len(rows), "bands": dict(Counter(l["band"] for l in rows)), "with_email": sum(1 for l in rows if l["primary"]["email"]),
                          "with_mobile": sum(1 for l in rows if l["primary"]["mobile"]), "with_backup": sum(1 for l in rows if l["backup"]["name"]),
                          "debtor_named": sum(1 for l in rows if l["debtor"]), "face": round(sum(float(l["face"]) for l in rows)),
                          "fund": round(sum(l["fund"] for l in rows)), "agio": round(sum(l["fund"] for l in rows) * (1 - DEFAULT_SHARE)),
                          "proof": dict(Counter(l["proof"].split(" ")[0] for l in rows)), "no_route": sum(1 for l in rows if "no route" in l["flags"].lower()),
                          "on_hold": sum(1 for l in rows if l["hold"]), "face_on_hold": round(sum(float(l["face"]) for l in rows if l["hold"]))}
    return {"trade": tally(by["TRADE"]), "financial": tally(by["FINANCIAL"]), "state": tally(by["STATE-OWNED"]), "book": tally(every),
            "fit": dict(Counter(l["fit"] for l in every)), "excluded": Counter(x["reason"].split(":")[0].split(" — ")[0].split(" (")[0] for x in excluded),
            "excluded_rows": [(x["razao"] or x["name_printed"], x["debtor"], float(x["face"]), x["reason"]) for x in excluded]}
