"""The call sheet — an XLSX workbook closers dial from.

Sheets: README · Assumptions (editable pricing inputs) · Call List (FLOOR band) · Pooling (100–200k)
· Sources (every document with its reconcile status) · Quarantine. Every lead row carries the
proof of claim (source link, page/row, value as printed, which list, publication date) and the
quote is an Excel formula off the Assumptions tab, so changing the IRR reprices the sheet.
"""
from __future__ import annotations

import json
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

STATUSES = ["NEW", "QUEUED", "EMAILED", "CALLED", "VOICEMAIL", "CONNECTED", "INTERESTED", "NOT INTERESTED", "DEAD", "DEAL"]
PLAN_STATUSES = ["CONFIRMED", "ESTIMATED", "UNKNOWN", "CONVERTED_TO_BANKRUPTCY"]
HEAD_FILL = PatternFill("solid", fgColor="1F3A5F")
CRM_FILL = PatternFill("solid", fgColor="FFF4D6")
PROOF_FILL = PatternFill("solid", fgColor="E8F1E8")
PRICE_FILL = PatternFill("solid", fgColor="E6EEF8")
THIN = Side(style="thin", color="C8C8C8")

COLS = [
    # (header, width, group)
    ("Rank", 6, "crm"), ("Status", 14, "crm"), ("Owner", 12, "crm"), ("Attempts", 9, "crm"), ("Last touch", 12, "crm"),
    ("Next action", 22, "crm"), ("Next date", 12, "crm"), ("Notes", 40, "crm"),
    ("Company (RFB)", 42, "lead"), ("Name as printed", 42, "lead"), ("CNPJ root", 11, "lead"), ("Establishment CNPJs", 22, "lead"),
    ("Class III face (R$)", 18, "lead"), ("Claims", 7, "lead"),
    ("Plan status", 14, "price"), ("Stage", 12, "price"), ("Case filed", 12, "price"), ("Years since filing", 9, "price"),
    ("Est. recovery %", 10, "price"), ("Years to payment (est.)", 10, "price"), ("Fund price @ IRR (R$)", 18, "price"),
    ("Fund price (cents)", 10, "price"), ("OUR QUOTE (R$)", 18, "price"), ("Quote (cents)", 10, "price"),
    ("Debtor", 40, "case"), ("Case number", 26, "case"), ("Court", 8, "case"), ("Administrator (site)", 26, "case"),
    ("Decision-maker", 34, "contact"), ("Role", 20, "contact"), ("Phone", 16, "contact"), ("Phone 2", 16, "contact"),
    ("Email", 30, "contact"), ("UF", 5, "contact"), ("City", 20, "contact"), ("CNAE", 40, "contact"), ("Porte", 12, "contact"),
    ("Situação", 10, "contact"),
    ("PROOF: source document", 60, "proof"), ("Page / row", 22, "proof"), ("Value as printed", 34, "proof"),
    ("Which list", 12, "proof"), ("Published", 12, "proof"), ("Reconciled", 14, "proof"), ("Data flags", 40, "proof"), ("Doc id", 14, "proof"),
]


import re as _re
_AJ_WORDS = _re.compile(r"administra[çc][ãa]o judicial|administradora? judicial|\bAJ\b|processos?|home|in[íi]cio|p[áa]gina|recupera[çc][õo]es judiciais|fal[êe]ncias?|consultoria|per[íi]cia|escrit[óo]rio|advogados|\.com\.br|\.adm\.br", _re.I)
_RJ_LEAD = _re.compile(r"(?:recupera[çc][ãa]o judicial|RJ|fal[êe]ncia)\s+(?:d[eao]s?\s+|do grupo\s+)?(.+)", _re.I)


_JUNK_DEBTOR = _re.compile(r"\bOAB\b|^DRA?\.|\bADVOGAD|\bDOUTOR|\bJUIZ|\bDESEMBARGADOR|\bMINIST[ÉE]RIO P[ÚU]BLICO|^(?:e\s+)?(?:na\s+|de\s+|da\s+)?fal[êe]ncia|\bvara\b|\bcomarca\b|\btribunal\b|\bju[íi]zo\b|\bempresarial\b|\bconcurso\b|\bcredores?\b|\bedital\b|\brela[çc][ãa]o\b|\blista\b|\bquadro\b|^(?:e|de|da|do|dos|das|na|no)\b", _re.I)
_DEBTOR_PATTERNS = [
    _re.compile(r"RECUPERANDA\(?S?\)?\s*[:\-–]\s*(.+?)(?=\s*(?:,\s*CNPJ|\s+CNPJ|\(|;|\n|$))", _re.I),
    _re.compile(r"\b(?:REQUERENTE|AUTOR(?:A|ES)?|DEVEDOR(?:A|ES)?|RECUPERANDAS?)\s*[:\-–]\s*(.+?)(?=\s*(?:\(|,\s*CNPJ|;|\n|$))", _re.I),
    _re.compile(r"([A-ZÀ-Ú0-9&][A-ZÀ-Ú0-9&.,\-'\s]{4,90}?)\s*\((?:REQUERENTE|AUTOR|RECUPERANDA)\)", _re.I),
    _re.compile(r"ajuizad[ao]\s+pel[ao]s?\s*:?\s*(?:\(i\)\s*)?(.+?)(?=\s*(?:\(|;|,|\n))", _re.I),
    _re.compile(r"RECUPERA[ÇC][ÃA]O JUDICIAL\s+(?:D[EAO]S?\s+)(?:EMPRESAS?\s+)?(GRUPO\s+[^\n,;(]{2,60}|[A-ZÀ-Ú0-9&][^\n,;(]{3,80}?(?:LTDA\.?|S\.?A\.?|S/A|EIRELI|\bME\b|\bEPP\b|CIA\.?|E OUTR[AO]S?))", _re.I),
    _re.compile(r"\b(?:RJ|RECUPERA[ÇC][ÃA]O JUDICIAL)\s+D[EAO]S?\s+(?:GRUPO\s+)?([A-ZÀ-Ú][A-ZÀ-Ú0-9&.\-'\s]{3,60}?)\s*(?:\n|,|;|\(|$)", _re.I),
]
_DEBTOR_CACHE = {}


def _debtor_from_pdf(db, doc_id, path):
    """First credible debtor name printed in the document's first pages; cached per document."""
    if doc_id in _DEBTOR_CACHE:
        return _DEBTOR_CACHE[doc_id]
    db.execute("CREATE TABLE IF NOT EXISTS doc_debtor(doc_id TEXT PRIMARY KEY, debtor TEXT)")
    row = db.execute("SELECT debtor FROM doc_debtor WHERE doc_id=?", (doc_id,)).fetchone()
    if row is not None:
        _DEBTOR_CACHE[doc_id] = row[0]
        return row[0]
    found = ""
    try:
        from ..extract.pdfio import open_pdf, page_text
        with open_pdf(path) as pdf:
            head = "\n".join(page_text(p) for p in pdf.pages[:2])
        for pat in _DEBTOR_PATTERNS:
            for m in pat.finditer(head):
                cand = _re.sub(r"\s+", " ", m.group(1)).strip(" -–—|:.,;")
                cand = _re.sub(r"\s*(?:\(|\bCNPJ\b).*$", "", cand).strip()
                if 3 <= len(cand) <= 90 and not _JUNK_DEBTOR.search(cand) and not _AJ_WORDS.search(cand) and _re.search(r"[A-Za-zÀ-ú]{3}", cand):
                    found = cand
                    break
            if found:
                break
    except Exception:  # noqa: BLE001
        found = ""
    db.execute("INSERT OR REPLACE INTO doc_debtor VALUES(?,?)", (doc_id, found))
    db.commit()
    _DEBTOR_CACHE[doc_id] = found
    return found


def _clean_debtor(hint, page_title, domain):
    """Best available debtor name: the case page title minus the administrator's branding, else the PDF hint."""
    cands = []
    for src in (page_title or "", hint or ""):
        for part in _re.split(r"\s+[–—|:]+\s+|\s+-\s+|\s+::\s+|\s+»\s+", src):
            part = part.strip(" -–—|:.")
            if not part or len(part) < 3:
                continue
            m = _RJ_LEAD.match(part)
            if m:
                part = m.group(1).strip(" -–—|:.")
            if _AJ_WORDS.search(part) or (domain and domain.split(".")[0].lower() in part.lower()):
                continue
            if _re.fullmatch(r"[\d\W]+", part):
                continue
            cands.append(part)
    cands = [c for c in cands if not _JUNK_DEBTOR.search(c)]
    if not cands:
        return ""
    # prefer a company-looking candidate, then the longest
    cands.sort(key=lambda c: (bool(_re.search(r"\b(LTDA|S\.?A\.?|S/A|EIRELI|ME|EPP|GRUPO|CIA)\b", c, _re.I)), len(c)), reverse=True)
    return cands[0][:80]


def _phone(p):
    """RFB phones are DDD+number digits: 1144634833 → (11) 4463-4833."""
    d = "".join(ch for ch in str(p or "") if ch.isdigit())
    if len(d) == 10:
        return f"({d[:2]}) {d[2:6]}-{d[6:]}"
    if len(d) == 11:
        return f"({d[:2]}) {d[2:7]}-{d[7:]}"
    return str(p or "")


def _plan_status(db, run_id, case_number, stage):
    if stage == "CONVERTED_TO_BANKRUPTCY":
        return "CONVERTED_TO_BANKRUPTCY"
    if not case_number:
        return "UNKNOWN"
    n = db.execute("SELECT COUNT(*) FROM documents WHERE run_id=? AND case_number=? AND doc_type IN ('PLAN','HOMOLOG')", (run_id, case_number)).fetchone()[0]
    if n:
        return "CONFIRMED"
    row = db.execute("SELECT filing_date, rj_granted_signal FROM cases WHERE case_number=?", (case_number,)).fetchone()
    return "ESTIMATED" if row and row[0] else "UNKNOWN"


def _lead_rows(db, run_id, band):
    q = db.execute("""
        SELECT t.cnpj_basico, t.case_number, t.doc_id, t.class_iii_face_sum, t.establishment_cnpjs, t.claim_count,
               t.creditor_name_as_printed, t.debtor_name, t.stage, t.flags,
               c.razao_social, c.phone, c.phone2, c.email, c.uf, c.municipio, c.cnae_desc, c.porte, c.situacao_desc,
               c.is_bank, c.is_public, c.is_inactive, c.rfb_source,
               d.source_url, d.doc_type, d.publication_date, d.status, d.administrator_hint,
               cs.court, cs.filing_date, rd.page_title, d.debtor_name_hint, d.file_path
        FROM targets t
        LEFT JOIN companies c ON c.cnpj_basico=t.cnpj_basico
        LEFT JOIN documents d ON d.doc_id=t.doc_id AND d.run_id=t.run_id
        LEFT JOIN raw_documents rd ON rd.sha1=t.doc_id
        LEFT JOIN cases cs ON cs.case_number=t.case_number
        WHERE t.run_id=? AND t.band=? ORDER BY CAST(t.class_iii_face_sum AS REAL) DESC""", (run_id, band)).fetchall()
    out = []
    for r in q:
        (root, case, doc_id, face, ests, ncl, name_printed, debtor, stage, flags, razao, phone, phone2, email, uf, city, cnae, porte, sit,
         is_bank, is_public, is_inactive, rfb_src, url, dtype, pub, dstatus, adm, court, filed, page_title, dhint, fpath) = r
        # one source of truth with the closer workbook: the verified/resolved display name in debtors; PDF/hint heuristics only as fallback
        resolved = db.execute("SELECT display_name FROM debtors WHERE case_number=?", (case,)).fetchone() if case else None
        debtor = (resolved[0] if resolved and resolved[0] else "") or (_debtor_from_pdf(db, doc_id, fpath) if fpath else "") or _clean_debtor(dhint or debtor, page_title, adm)
        if is_bank or is_public or is_inactive or "LIKELY_FINANCIAL_BY_NAME" in (flags or "") or "LIKELY_PUBLIC_BY_NAME" in (flags or ""):
            continue   # excluded by step G; kept in targets.csv with flags
        dm = db.execute("SELECT person_name, role FROM contacts WHERE cnpj_basico=? ORDER BY CASE role_code WHEN 49 THEN 0 WHEN 5 THEN 1 WHEN 16 THEN 2 WHEN 10 THEN 3 ELSE 9 END LIMIT 1", (root,)).fetchone()
        claims = db.execute("SELECT page, row_index, value_as_printed FROM claims WHERE doc_id=? AND run_id=? AND document_number IN (%s) AND class='III' AND currency='BRL' ORDER BY page, row_index"
                            % ",".join("?" * len((ests or "").split("|"))), (doc_id, run_id, *((ests or "").split("|")))).fetchall()
        out.append({
            "root": root, "case": case, "doc_id": doc_id, "face": Decimal(face), "ests": ests, "n": ncl, "name_printed": name_printed,
            "debtor": debtor, "stage": stage or "", "flags": flags, "razao": razao, "phone": phone, "phone2": phone2, "email": email,
            "uf": uf, "city": city, "cnae": cnae, "porte": porte, "sit": sit, "url": url, "dtype": dtype, "pub": pub, "dstatus": dstatus,
            "adm": adm, "court": court, "filed": filed, "dm": dm, "claims": claims,
            "plan_status": _plan_status(db, run_id, case, stage),
        })
    return out


def _write_lead_sheet(ws, leads, title):
    ws.title = title
    ws.freeze_panes = "J2"
    for j, (h, w, g) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = HEAD_FILL; c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 32
    dv = DataValidation(type="list", formula1='"' + ",".join(STATUSES) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    for i, L in enumerate(leads, start=2):
        row = i
        face_cell = f"M{row}"; ps = f"O{row}"; ysf = f"R{row}"
        vals = [
            i - 1, "NEW", "", 0, "", "", "", "",
            L["razao"] or "", L["name_printed"], L["root"], (L["ests"] or "").replace("|", ", "),
            float(L["face"]), L["n"],
            L["plan_status"], L["stage"], (date.fromisoformat(L["filed"]) if L["filed"] and len(L["filed"]) == 10 else ""),
            f'=IF(Q{row}="","",ROUND((TODAY()-Q{row})/365.25,1))',
            f"=VLOOKUP({ps},Assumptions!$A$7:$E$10,2,FALSE)",
            f'=MAX(0.5,VLOOKUP({ps},Assumptions!$A$7:$E$10,3,FALSE)+VLOOKUP({ps},Assumptions!$A$7:$E$10,4,FALSE)/2-IF({ysf}="",0,{ysf}))',
            f"={face_cell}*S{row}/(1+Assumptions!$B$3)^T{row}",
            f"=IF({face_cell}=0,\"\",ROUND(U{row}/{face_cell}*100,1))",
            f"=U{row}*Assumptions!$B$4",
            f"=IF({face_cell}=0,\"\",ROUND(W{row}/{face_cell}*100,1))",
            L["debtor"] or "", L["case"] or "(not found in document)", L["court"] or "", L["adm"] or "",
            (L["dm"][0] if L["dm"] else ""), (L["dm"][1] if L["dm"] else ""), _phone(L["phone"]), _phone(L["phone2"]), L["email"] or "",
            L["uf"] or "", L["city"] or "", L["cnae"] or "", L["porte"] or "", L["sit"] or "",
            L["url"] or "", "; ".join(f"p{p} r{r}" for p, r, _ in L["claims"]),
            "; ".join(f"p{p}: R$ {v}" for p, r, v in L["claims"]),
            L["dtype"] or "", L["pub"] or "", L["dstatus"] or "", (L["flags"] or "").replace("|", " "), (L["doc_id"] or "")[:12],
        ]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=j, value=v)
            g = COLS[j - 1][2]
            if g == "crm":
                c.fill = CRM_FILL
            elif g == "proof":
                c.fill = PROOF_FILL
            elif g == "price":
                c.fill = PRICE_FILL
        ws.cell(row=row, column=13).number_format = '#,##0.00'
        for col in (21, 23):
            ws.cell(row=row, column=col).number_format = '#,##0'
        ws.cell(row=row, column=19).number_format = '0%'
        ws.cell(row=row, column=20).number_format = '0.0'
        ws.cell(row=row, column=17).number_format = 'yyyy-mm-dd'
        if L["url"]:
            c = ws.cell(row=row, column=39); c.hyperlink = L["url"]; c.font = Font(color="0563C1", underline="single")
        dv.add(f"B{row}")
        for j in (23,):
            ws.cell(row=row, column=j).font = Font(bold=True)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{max(2, len(leads) + 1)}"


def _assumptions(ws):
    ws.title = "Assumptions"
    ws.column_dimensions["A"].width = 30; ws.column_dimensions["F"].width = 70
    ws["A1"] = "PRICING ASSUMPTIONS — edit the yellow cells; every quote on the call sheet recalculates"; ws["A1"].font = Font(bold=True, size=12)
    ws["A3"] = "Target investor IRR (annual)"; ws["B3"] = 0.25; ws["B3"].number_format = "0%"
    ws["A4"] = "Buy price as share of fund price (we pay 60% → 0.60)"; ws["B4"] = 0.60; ws["B4"].number_format = "0%"
    ws["F3"] = "Fund price = face × recovery% ÷ (1+IRR)^(years to payment). A fund needing 25% IRR on a plan paying 30% over years 3–10 pays ~8 cents, not 25 — check the 7 buyers' real numbers against this."
    ws["F4"] = "Our quote = fund price × this share. The doc rule is 60% of mandate price; expected 55–70% once Gate 6 data exists."
    ws["A6"] = "Plan status"; ws["B6"] = "Class III recovery % of face"; ws["C6"] = "Grace (years)"; ws["D6"] = "Term (years)"; ws["E6"] = "Note"
    for c in "ABCDE":
        ws[f"{c}6"].font = Font(bold=True, color="FFFFFF"); ws[f"{c}6"].fill = HEAD_FILL
    rows = [
        ("CONFIRMED", 0.40, 2, 8, "A plan/homologation document exists for the case. ASSUMED terms until plan parsing lands — replace with the plan's own class III terms."),
        ("ESTIMATED", 0.30, 2, 8, "No plan document found; case is live in DataJud. ASSUMED typical Brazilian class III terms."),
        ("UNKNOWN", 0.25, 3, 8, "Case not matched to DataJud. ASSUMED, conservative."),
        ("CONVERTED_TO_BANKRUPTCY", 0.05, 4, 4, "Falência decreed. Class III ranks behind labour and secured; assume near-zero."),
    ]
    for i, (a, b, c, d, e) in enumerate(rows, start=7):
        ws[f"A{i}"] = a; ws[f"B{i}"] = b; ws[f"C{i}"] = c; ws[f"D{i}"] = d; ws[f"E{i}"] = e
        ws[f"B{i}"].number_format = "0%"
        for col in "BCD":
            ws[f"{col}{i}"].fill = PatternFill("solid", fgColor="FFF2CC")
    for col in ("B3", "B4"):
        ws[col].fill = PatternFill("solid", fgColor="FFF2CC")
    ws["A13"] = "Everything in this tab is an ASSUMPTION, not court data. The face value on the call sheet IS court data — it links to the page."
    ws["A13"].font = Font(italic=True)


def _readme(ws, run_id, counts):
    ws.title = "README"
    ws.column_dimensions["A"].width = 120
    lines = [
        ("ANTEPARO — CALL SHEET", True),
        (f"Run {run_id} · generated {date.today().isoformat()} · {counts['floor']} leads ≥ R$200k · {counts['pool']} pooling leads (R$100–200k) · {counts['docs']} source documents · {counts['cases']} cases", False),
        ("", False),
        ("WHAT EACH ROW IS", True),
        ("One company × one recovery case. The company holds class III (trade) credit against the debtor, admitted in the court's own creditor list. All establishments of the company (same 8-digit CNPJ root) are summed.", False),
        ("", False),
        ("THE NUMBER", True),
        ("'Class III face' is the admitted value AS PRINTED in the court document — what the company is OWED, updated only to the filing date (art. 9 II). It is NOT what will be paid.", False),
        ("'OUR QUOTE' is a model: face → assumed plan terms → discounted at the investor IRR → fund price → our share. Edit the Assumptions tab. It is not court data.", False),
        ("", False),
        ("PROOF OF CLAIM (green columns)", True),
        ("Click 'PROOF: source document', go to the page/row listed, and read the value as printed. If a seller asks 'how did you get this?', that link is the answer.", False),
        ("'Reconciled' = OK means the document's extracted class totals matched its own printed totals within 0.5%. OK_NO_TOTALS means the document prints no totals — the number was read from its own row but could not be cross-checked; quote with that caveat.", False),
        ("", False),
        ("DECISION-MAKER", True),
        ("From the Receita Federal partner register (sócio-administrador / administrador / diretor / presidente). Phone and email are the company's registry contact. Confidence LOW = registry only; nothing here was guessed or looked up from memory.", False),
        ("", False),
        ("CRM COLUMNS (yellow)", True),
        ("Status (dropdown) · Owner · Attempts · Last touch · Next action · Next date · Notes. Sort/filter freely; the proof and quote columns travel with the row.", False),
        ("", False),
        ("EXCLUDED FROM THIS SHEET", True),
        ("Individuals (CPF), banks (CNAE 64–66), public entities, inactive companies, foreign-currency claims, extraconcursal sections, quarantined documents. They remain in the CSV tables with flags.", False),
    ]
    for i, (t, b) in enumerate(lines, 1):
        ws.cell(row=i, column=1, value=t).font = Font(bold=b, size=12 if b else 11)
        ws.cell(row=i, column=1).alignment = Alignment(wrap_text=True)


def _sources(ws, db, run_id):
    ws.title = "Sources"
    hdr = ["Doc id", "Case number", "Type", "Status", "Pages", "Strategy", "Layout", "Reconcile checks", "Published", "Debtor (hint)", "Administrator", "URL"]
    for j, h in enumerate(hdr, 1):
        c = ws.cell(row=1, column=j, value=h); c.font = Font(bold=True, color="FFFFFF"); c.fill = HEAD_FILL
    widths = [14, 26, 12, 14, 7, 10, 14, 60, 12, 40, 26, 70]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    rows = db.execute("SELECT doc_id, case_number, doc_type, status, pages, strategy, layout_id, reconcile_json, publication_date, debtor_name_hint, administrator_hint, source_url FROM documents WHERE run_id=? ORDER BY status, doc_type", (run_id,)).fetchall()
    for i, r in enumerate(rows, 2):
        checks = ""
        try:
            rec = json.loads(r[7] or "{}")
            checks = " · ".join(f"{c['class']}: {'PASS' if c['pass'] else 'FAIL'} Δ{c.get('delta','')}" for c in rec.get("checks", []))
        except Exception:  # noqa: BLE001
            pass
        vals = [r[0][:12], r[1], r[2], r[3], r[4], r[5], r[6], checks, r[8], r[9], r[10], r[11]]
        for j, v in enumerate(vals, 1):
            ws.cell(row=i, column=j, value=v)
        if r[11]:
            c = ws.cell(row=i, column=12); c.hyperlink = r[11]; c.font = Font(color="0563C1", underline="single")
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:L{max(2, len(rows) + 1)}"


def _needs_cnpj(ws, db, run_id):
    """Class III rows above the floor whose list prints no CNPJ — leads that need name resolution."""
    from ..steps.build import best_document_per_case, FLOOR
    ws.title = "Needs CNPJ (name only)"
    hdr = ["Name as printed", "Class III value (R$)", "Debtor", "Case number", "Which list", "Reconciled", "PROOF: source document", "Page / row", "Flags"]
    for j, h in enumerate(hdr, 1):
        c = ws.cell(row=1, column=j, value=h); c.font = Font(bold=True, color="FFFFFF"); c.fill = HEAD_FILL
    for j, w in enumerate([48, 18, 40, 26, 12, 14, 60, 12, 40], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    i = 2
    for case_key, (doc_id, dtype, status) in best_document_per_case(db, run_id).items():
        rows = db.execute("""SELECT c.creditor_name_as_printed, c.value_brl, d.debtor_name_hint, d.case_number, d.doc_type, d.status, d.source_url, c.page, c.row_index, c.flags
                             FROM claims c JOIN documents d ON d.doc_id=c.doc_id AND d.run_id=c.run_id
                             WHERE c.doc_id=? AND c.run_id=? AND c.class='III' AND c.currency='BRL' AND c.document_type='NONE'
                             AND c.value_brl IS NOT NULL AND CAST(c.value_brl AS REAL) >= ? AND c.flags NOT LIKE '%NOT_A_CLAIM%'
                             ORDER BY CAST(c.value_brl AS REAL) DESC""", (doc_id, run_id, float(FLOOR))).fetchall()
        for (name, val, debtor, case, dt, st, url, page, ridx, flags) in rows:
            vals = [name, float(val), debtor or "", case or "", dt, st, url or "", f"p{page} r{ridx}", (flags or "").replace("|", " ")]
            for j, v in enumerate(vals, 1):
                ws.cell(row=i, column=j, value=v)
            ws.cell(row=i, column=2).number_format = '#,##0.00'
            if url:
                c = ws.cell(row=i, column=7); c.hyperlink = url; c.font = Font(color="0563C1", underline="single")
            i += 1
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:I{max(2, i - 1)}"
    return i - 2


def _diff_tab(ws, db, run_id, old_csv):
    """Old Drive sheet vs this run, by CNPJ root × case: every changed number, side by side."""
    import csv
    from pathlib import Path
    ws.title = "Diff vs Drive sheet"
    hdr = ["Creditor (old sheet)", "CNPJ (old)", "Case", "OLD amount (R$)", "NEW class III face, root sum (R$)", "Delta (R$)", "Verdict", "Old status", "Why"]
    for j, h in enumerate(hdr, 1):
        c = ws.cell(row=1, column=j, value=h); c.font = Font(bold=True, color="FFFFFF"); c.fill = HEAD_FILL
    for j, w in enumerate([44, 16, 26, 18, 22, 18, 16, 12, 60], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    if not Path(old_csv).exists():
        ws.cell(row=2, column=1, value="old sheet not available")
        return 0
    new = {}
    for root, case, face, name, ests, n, flags in db.execute("SELECT cnpj_basico, case_number, class_iii_face_sum, creditor_name_as_printed, establishment_cnpjs, claim_count, flags FROM targets WHERE run_id=?", (run_id,)):
        new[(root, case or "")] = (Decimal(face), name, ests, n, flags)
    i = 2
    with open(old_csv, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            key = (r["cnpj"][:8], r["case_number"])
            try:
                old_amt = Decimal(r["amount_old"]) if r["amount_old"] else None
            except Exception:  # noqa: BLE001
                old_amt = None
            if key in new:
                new_amt, nm, ests, n, flags = new[key]
                delta = (new_amt - old_amt) if old_amt is not None else None
                if delta is None:
                    verdict, why = "NEW_ONLY", "old sheet had no amount"
                elif abs(delta) < 1:
                    verdict, why = "SAME", ""
                else:
                    verdict = "CHANGED"
                    why = []
                    if n and n > 1:
                        why.append(f"{n} claims summed on the 8-digit root ({(ests or '').count('|') + 1} establishments)")
                    if old_amt and new_amt and new_amt / old_amt > Decimal(5):
                        why.append("likely dropped leading digit in old parse")
                    why = "; ".join(why) or "value re-read from the document's own row"
                vals = [r["creditor_name"], r["cnpj"], r["case_number"], float(old_amt) if old_amt is not None else "", float(new_amt), float(delta) if delta is not None else "", verdict, r["status_old"], why]
            else:
                vals = [r["creditor_name"], r["cnpj"], r["case_number"], float(old_amt) if old_amt is not None else "", "", "", "NOT_IN_NEW_RUN", r["status_old"], "case/document not harvested or below floor in this run"]
            for j, v in enumerate(vals, 1):
                c = ws.cell(row=i, column=j, value=v)
                if j in (4, 5, 6) and isinstance(v, float):
                    c.number_format = '#,##0.00'
            if vals[6] == "CHANGED":
                for j in range(1, 10):
                    ws.cell(row=i, column=j).fill = PatternFill("solid", fgColor="FCE4D6")
            i += 1
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:I{max(2, i - 1)}"
    return i - 2


def export_xlsx(db, run_id, path: str, old_csv: str | None = None):
    wb = Workbook()
    floor = _lead_rows(db, run_id, "FLOOR")
    pool = _lead_rows(db, run_id, "POOL")
    counts = {"floor": len(floor), "pool": len(pool),
              "docs": db.execute("SELECT COUNT(*) FROM documents WHERE run_id=?", (run_id,)).fetchone()[0],
              "cases": db.execute("SELECT COUNT(DISTINCT case_number) FROM documents WHERE run_id=? AND case_number IS NOT NULL", (run_id,)).fetchone()[0]}
    _readme(wb.active, run_id, counts)
    _assumptions(wb.create_sheet())
    _write_lead_sheet(wb.create_sheet(), floor, "Call List")
    _write_lead_sheet(wb.create_sheet(), pool, "Pooling (100-200k)")
    counts["needs_cnpj"] = _needs_cnpj(wb.create_sheet(), db, run_id)
    _sources(wb.create_sheet(), db, run_id)
    if old_csv:
        counts["diff_rows"] = _diff_tab(wb.create_sheet(), db, run_id, old_csv)
    wb.save(path)
    return counts
